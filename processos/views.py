from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from .forms import ProcessoForm, AndamentoForm, ComentarioProcessoForm
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import datetime, timedelta
from calendar import month_name
from .models import Processo, Fase, Status, Camara, Tipo, Especie, TarefaDoDia, ProcessoAndamento, Fase, Resultado, Tema, ComentarioProcesso
import locale
from calendar import month_name
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.utils.timezone import make_aware
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from .metrics import get_advanced_metrics
from accounts.models import UserProfile
from django.urls import reverse
from django.utils.timezone import now
from django.views import View
from django.db.models import Subquery, OuterRef
from django.contrib import messages
from rest_framework import generics
from . import models, forms, serializers
import pandas as pd
from django.contrib.auth.models import User
from django.utils import timezone
from datetime import datetime
from django.shortcuts import render
from django.contrib import messages
from django.http import HttpResponseRedirect
from .models import Processo, Especie, Camara, Tema
from django.db import DatabaseError

from django.views.decorators.http import require_http_methods
from .models import Processo, MetaSemanal, Aviso

    
from django.http import HttpResponseForbidden

def is_gestor(user):

    return user.groups.filter(name='Gestor(a)').exists()


@login_required
@require_http_methods(["POST"])
def adicionar_processo_meta(request, processo_id):

    try:
        # Verificar se usuário é gestor
        if not is_gestor(request.user):
            return JsonResponse({
                'success': False, 
                'message': 'Apenas gestores podem gerenciar metas semanais.'
            }, status=403)

        processo = get_object_or_404(Processo, id=processo_id)
        
        # Verificar se o processo tem responsável
        if not processo.usuario:
            return JsonResponse({
                'success': False, 
                'message': 'Este processo não possui responsável definido.'
            }, status=400)

        # Calcular semana atual
        hoje = timezone.localdate()
        inicio_semana = hoje - timedelta(days=hoje.weekday())
        fim_semana = inicio_semana + timedelta(days=6)

        # Buscar ou criar meta semanal do responsável pelo processo
        meta, created = MetaSemanal.objects.get_or_create(
            usuario=processo.usuario,  # Meta do responsável, não do gestor
            semana_inicio=inicio_semana,
            semana_fim=fim_semana,
            defaults={'meta_qtd': 0}
        )

        # Verificar se o processo já está na meta
        if meta.processos.filter(id=processo_id).exists():
            return JsonResponse({
                'success': False,
                'message': f'Este processo já está na meta de {processo.usuario.get_full_name()}.'
            })

        # Adicionar processo à meta
        meta.processos.add(processo)
        meta.meta_qtd = meta.processos.count()
        meta.save()

        return JsonResponse({
            'success': True,
            'message': f'Processo {processo.numero_processo} adicionado à meta de {processo.usuario.get_full_name()}!',
            'meta_qtd': meta.meta_qtd,
            'responsavel': processo.usuario.get_full_name()
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao adicionar processo à meta: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def remover_processo_meta(request, processo_id):

    try:
        # Verificar se usuário é gestor
        if not is_gestor(request.user):
            return JsonResponse({
                'success': False, 
                'message': 'Apenas gestores podem gerenciar metas semanais.'
            }, status=403)

        processo = get_object_or_404(Processo, id=processo_id)
        
        # Verificar se o processo tem responsável
        if not processo.usuario:
            return JsonResponse({
                'success': False, 
                'message': 'Este processo não possui responsável definido.'
            }, status=400)

        # Calcular semana atual
        hoje = timezone.localdate()
        inicio_semana = hoje - timedelta(days=hoje.weekday())
        fim_semana = inicio_semana + timedelta(days=6)

        # Buscar meta semanal do responsável
        try:
            meta = MetaSemanal.objects.get(
                usuario=processo.usuario, 
                semana_inicio=inicio_semana,
                semana_fim=fim_semana
            )
        except MetaSemanal.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'{processo.usuario.get_full_name()} não possui meta semanal configurada.'
            })

        # Verificar se o processo está na meta
        if not meta.processos.filter(id=processo_id).exists():
            return JsonResponse({
                'success': False,
                'message': f'Este processo não está na meta de {processo.usuario.get_full_name()}.'
            })

        # Remover processo da meta
        meta.processos.remove(processo)
        meta.meta_qtd = meta.processos.count()
        meta.save()

        return JsonResponse({
            'success': True,
            'message': f'Processo {processo.numero_processo} removido da meta de {processo.usuario.get_full_name()}!',
            'meta_qtd': meta.meta_qtd,
            'responsavel': processo.usuario.get_full_name()
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao remover processo da meta: {str(e)}'
        }, status=500)


@login_required
def status_meta_semanal(request, user_id=None):
    """
    Retorna informações sobre a meta semanal
    Gestores podem ver qualquer meta, usuários comuns apenas a própria
    """
    try:

        if user_id and is_gestor(request.user):
            # Gestor consultando meta de outro usuário
            target_user = get_object_or_404(User, id=user_id)
        else:
            # Usuário consultando própria meta
            target_user = request.user

        # Calcular semana atual
        hoje = timezone.localdate()
        inicio_semana = hoje - timedelta(days=hoje.weekday())
        fim_semana = inicio_semana + timedelta(days=6)

        # Buscar meta semanal
        try:
            meta = MetaSemanal.objects.get(
                usuario=target_user,
                semana_inicio=inicio_semana,
                semana_fim=fim_semana
            )
            
            processos_na_meta = meta.processos.all()
            processos_concluidos = processos_na_meta.filter(concluido=True).count()
            
            return JsonResponse({
                'success': True,
                'meta': {
                    'id': meta.id,
                    'usuario': target_user.get_full_name(),
                    'meta_qtd': meta.meta_qtd,
                    'processos_total': processos_na_meta.count(),
                    'processos_concluidos': processos_concluidos,
                    'processos_pendentes': processos_na_meta.count() - processos_concluidos,
                    'percentual_conclusao': round((processos_concluidos / processos_na_meta.count() * 100), 1) if processos_na_meta.count() > 0 else 0,
                    'semana_inicio': inicio_semana.strftime('%d/%m/%Y'),
                    'semana_fim': fim_semana.strftime('%d/%m/%Y'),
                }
            })
            
        except MetaSemanal.DoesNotExist:
            return JsonResponse({
                'success': True,
                'meta': None,
                'message': f'Nenhuma meta semanal configurada para {target_user.get_full_name()} nesta semana.'
            })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao buscar status da meta: {str(e)}'
},status=500)

@login_required
@require_http_methods(["GET", "POST"])
def configurar_meta_semanal(request):
    if request.method == "POST":
        usuario_id = request.POST.get('usuario_id')
        processo_ids = request.POST.getlist('processo_ids[]')
        meta_qtd = request.POST.get('meta_qtd')
        periodo = request.POST.get('periodo')  # Período selecionado (ex.: "atual" ou "seguinte")

        # Validações
        if not usuario_id:
            return JsonResponse({'success': False, 'message': 'Usuário é obrigatório.'}, status=400)
        if not processo_ids or not meta_qtd:
            return JsonResponse({'success': False, 'message': 'Selecione pelo menos um processo.'}, status=400)
        if not periodo:
            return JsonResponse({'success': False, 'message': 'Período é obrigatório.'}, status=400)

        try:
            meta_qtd = int(meta_qtd)
            if meta_qtd <= 0:
                raise ValueError
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Quantidade de meta inválida.'}, status=400)

        # Calcular semanas dinamicamente
        hoje = timezone.localdate()
        inicio_semana_atual = hoje - timedelta(days=hoje.weekday())  # Segunda-feira da semana atual
        fim_semana_atual = inicio_semana_atual + timedelta(days=6)  # Domingo da semana atual
        inicio_semana_seguinte = inicio_semana_atual + timedelta(days=7)  # Segunda-feira da semana seguinte
        fim_semana_seguinte = inicio_semana_seguinte + timedelta(days=6)  # Domingo da semana seguinte

        # Determinar as datas do período selecionado
        if periodo == "atual":
            inicio_semana = inicio_semana_atual
            fim_semana = fim_semana_atual
        elif periodo == "seguinte":
            inicio_semana = inicio_semana_seguinte
            fim_semana = fim_semana_seguinte
        else:
            return JsonResponse({'success': False, 'message': 'Período inválido.'}, status=400)

        usuario = get_object_or_404(User, id=usuario_id)
        processos = Processo.objects.filter(id__in=processo_ids, usuario=usuario, concluido=False)

        # Cria ou atualiza a meta para o período selecionado
        meta, created = MetaSemanal.objects.get_or_create(
            usuario=usuario,
            semana_inicio=inicio_semana,
            semana_fim=fim_semana,
            defaults={'meta_qtd': meta_qtd}
        )
        if not created:
            meta.meta_qtd = meta_qtd
            meta.save()
        meta.processos.set(processos)

        return JsonResponse({'success': True, 'message': 'Meta salva com sucesso!'})

    # GET: Carregar usuários, processos e períodos disponíveis
    usuarios = User.objects.all()
    processos = Processo.objects.filter(concluido=False).select_related('especie', 'usuario')

    # Subquery para fase atual
    ultima_fase_subquery = ProcessoAndamento.objects.filter(
        processo=OuterRef('pk')
    ).order_by('-dt_criacao').values('fase__fase')[:1]

    # Adiciona anotação de fase atual
    processos = processos.annotate(fase_atual=Subquery(ultima_fase_subquery))

    # Monta dicionário agrupado por usuário
    processos_por_usuario = {}
    for user in usuarios:
        user_procs = processos.filter(usuario=user)
        processos_por_usuario[user.id] = [
            {
                'id': proc.id,
                'numero_processo': proc.numero_processo,
                'especie': proc.especie.especie if proc.especie else 'N/A',
                'tipo': proc.tipo.tipo if proc.tipo else 'Não informado', # Adicionado aqui
                'dias_no_gabinete': (timezone.now().date() - proc.antigo.date()).days if proc.antigo else 0,
                'fase_atual': proc.fase_atual or '—'
            }
            for proc in user_procs
        ]

    # Determina os períodos disponíveis (semana atual e semana seguinte)
    hoje = timezone.localdate()
    inicio_semana_atual = hoje - timedelta(days=hoje.weekday())
    fim_semana_atual = inicio_semana_atual + timedelta(days=6)
    inicio_semana_seguinte = inicio_semana_atual + timedelta(days=7)
    fim_semana_seguinte = inicio_semana_seguinte + timedelta(days=6)

    periodos = [
        {
            'label': f"Semana Atual ({inicio_semana_atual.strftime('%d/%m/%Y')} a {fim_semana_atual.strftime('%d/%m/%Y')})",
            'value': 'atual'
        },
        {
            'label': f"Semana Seguinte ({inicio_semana_seguinte.strftime('%d/%m/%Y')} a {fim_semana_seguinte.strftime('%d/%m/%Y')})",
            'value': 'seguinte'
        },
    ]

    return render(request, 'configurar_meta_semanal.html', {
        'usuarios': usuarios,
        'processos_por_usuario': processos_por_usuario,
        'periodos': periodos,
    })

from django.contrib.auth import get_user_model
from django.db.models import Prefetch, Count, Q, Avg
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


# Mapeamento de siglas para nomes completos
MAPEAMENTO_ESPECIES = {
    'ApCrim': 'Apelação Criminal',
    'ArRCrim': 'Agravo Regimental Criminal',
    'AI': 'Agravo de Instrumento',
    'AgExPe': 'Agravo em Execução Penal',
    'AgRCiv': 'Agravo Regimental Cível',
    'ApCiv': 'Apelação Cível',
    'ApCiv/RNCi': 'Apelação/Remessa Necessária Cível',
    'AC': 'Ação Cautelar',
    'ACP': 'Ação Civil Pública',
    'AIJE': 'Ação de Investigação Judicial Eleitoral',
    'ADJ': 'Ação Declaratória de Inexistência de Débito',
    'ADIN': 'Ação Direta de Inconstitucionalidade',
    'AR': 'Ação Rescisória',
    'CT': 'Carta Testemunhável',
    'CIC': 'Cautelar Inominada Criminal',
    'CC': 'Conflito de Competência',
    'CCCiv': 'Conflito de Competência Cível',
    'Desafor': 'Desaforamento',
    'ED': 'Embargos de Declaração',
    'EDCrim': 'Embargos de Declaração Criminal',
    'ExecFiscal': 'Execução Fiscal',
    'IDC': 'Incidente de Deslocamento de Competência',
    'ISM': 'Incidente de Sanidade Mental',
    'LIM': 'Liminar',
    'MSCrim': 'Mandado de Segurança Criminal',
    'MSCiv': 'Mandado de Segurança Cível',
    'PESE': 'Pedido de Efeito Suspensivo à Apelação',
    'PetCrim': 'Petição Criminal',
    'Rcl': 'Reclamação',
    'Resp': 'Recurso Especial',
    'RE': 'Recurso Extraordinário',
    'ROC': 'Recurso Ordinário Constitucional',
    'RSE': 'Recurso em Sentido Estrito',
    'RINT': 'Representação por Interceptação Telefônica',
    'RQD': 'Representação por Quebra de Domicílio',
    'HCCrim': 'Habeas Corpus Criminal',
    'RECURSO': 'Recurso Ordinário',
    'RevCrim': 'Revisão Criminal',
}

# Mapeamento de tags para nomes de usuário
MAPEAMENTO_TAG_USUARIO = {
    'Ass-Priscilla': 'priscillalessi',
    'Ass-Thais': 'thaiscamila',
    'Ass-Caio': 'caiocesar',
    'Ass-Izabela': 'izabelaortiz',
    'ASS-IZABELA': 'izabelaortiz',
    'Ass-Juliana': 'julianalodi',
    'Ass-Jurandy': 'jurandysilva',
    'Ass-Karen': 'karenaquino',
    'Ass-Manoel': 'manoellima',
    'Ass-Marco Antônio': 'marcoantonio',
    'ASS-Marco Antônio': 'marcoantonio',
    'Ass-Mirelli': 'mirellisilva',
    'Ass-Viviane': 'vivianelima',
    'Ass-Paulo': 'ass-paulo',
    'Ass-Ricardo': 'ricardo'
}

def importar_processos_view(request):
    if request.method == 'POST':
        if 'arquivo' not in request.FILES:
            messages.error(request, 'Nenhum arquivo selecionado.')
            return render(request, 'importar_processos.html')

        arquivo = request.FILES['arquivo']
        if not arquivo.name.endswith('.csv'):
            messages.error(request, 'O arquivo deve ser um CSV (.csv).')
            return render(request, 'importar_processos.html')

        try:
            # Forçar que a coluna 'prioridade' seja lida como string
            df = pd.read_csv(arquivo, sep=';', encoding='utf-8', dtype={'prioridade': str})


            # Verificar colunas esperadas
            colunas_esperadas = ['numeroProcesso', 'classeJudicial', 'assuntoPrincipal', 'tagsProcessoList', 'dataChegada', 'prioridade']
            colunas_faltando = [col for col in colunas_esperadas if col not in df.columns]
            if colunas_faltando:
                messages.error(request, f"Colunas faltando no CSV: {', '.join(colunas_faltando)}")
                return render(request, 'importar_processos.html')

            # Carregar usuários mapeados para evitar consultas repetidas
            usuarios = {u.username: u for u in User.objects.filter(username__in=MAPEAMENTO_TAG_USUARIO.values())}
            usuario_default = 'admin'
            if usuario_default not in usuarios and not User.objects.filter(username=usuario_default).exists():
                messages.error(request, f"Usuário padrão {usuario_default} não encontrado no banco de dados.")
                return render(request, 'importar_processos.html')

            processos_inseridos = 0
            processos_ignorados = 0
            try:
                tipo_monocratica = Tipo.objects.get(tipo="Monocrática")
            except Tipo.DoesNotExist:
                messages.error(request, "Erro crítico: O tipo 'Monocrática' não existe no banco de dados. Cadastre-o e tente novamente.")
                return render(request, 'importar_processos.html')
            except Exception as e:
                 messages.error(request, f"Erro ao buscar o tipo 'Monocrática': {str(e)}")
                 return render(request, 'importar_processos.html')

            processos_inseridos = 0
            processos_ignorados = 0

            for index, row in df.iterrows():
                try:
                    numero_processo = row['numeroProcesso']

                    # Ignorar processos já existentes e não concluídos
                    #processo_existente = Processo.objects.filter(numero_processo=numero_processo, concluido=False).first()
                    #if processo_existente:
                    #    processos_ignorados += 1
                    #    continue

                    # Processar espécie
                    especie = None
                    if pd.notna(row.get('classeJudicial')) and row['classeJudicial'].strip():
                        sigla = str(row['classeJudicial']).strip().upper()
                        nome_especie = MAPEAMENTO_ESPECIES.get(sigla, sigla)  # Usar sigla como fallback
                        especie, _ = Especie.objects.get_or_create(
                            sigla=sigla,
                            defaults={
                                'especie': nome_especie,
                                'dt_criacao': timezone.now(),
                                'dt_atualizacao': timezone.now(),
                            }
                        )
                    else:
                        print(f"Processo {numero_processo} sem classe judicial, ignorando espécie.")

                    # Processar tema
                    tema = None
                    if pd.notna(row.get('assuntoPrincipal')) and row['assuntoPrincipal'].strip():
                        tema, _ = Tema.objects.get_or_create(
                            nome=row['assuntoPrincipal'],
                            defaults={
                                'nome': row['assuntoPrincipal'],
                                'dt_criacao': timezone.now(),
                                'dt_atualizacao': timezone.now(),
                                'ativo': True
                            }
                        )

                      # Garantir que temos um valor limpo para nomeTarefa
                    nome_tarefa_limpo = ""
                    if pd.notna(row.get('nomeTarefa')):
                        nome_tarefa_limpo = row['nomeTarefa'].strip() # Pega o valor e remove espaços

                    # Marcar despacho como True (usando a nova variável limpa)
                    despacho = False
                    if 'Minutar despacho ou decisão' in nome_tarefa_limpo.lower():
                        despacho = True

                    # Processar o Tipo (Monocrática) (usando a mesma variável limpa)
                    tipo_processo = None  # Inicia como nulo por padrão
                    if nome_tarefa_limpo == "Minutar decisão monocrática":
                        tipo_processo = tipo_monocratica

                    # Processar usuário e tags de matéria a partir das tags
                    usuario = None
                    tags_materia = None
                    if pd.notna(row.get('tagsProcessoList')) and row['tagsProcessoList'].strip():
                        todas_tags = [tag.strip() for tag in str(row['tagsProcessoList']).split(',') if tag.strip()]
                        # Tags Ass- para usuário
                        tags_ass = [t for t in todas_tags if t.startswith('Ass-')]
                        # Tags de matéria: ignorar Ass-, Minutado, Resumo Feito
                        tags_mat = [t for t in todas_tags if not t.startswith('Ass-') and t not in ('Minutado', 'Resumo Feito')]
                        if tags_mat:
                            tags_materia = ', '.join(tags_mat)
                        usuario_encontrado = False
                        for tag in tags_ass:
                            username_mapeado = MAPEAMENTO_TAG_USUARIO.get(tag)
                            if username_mapeado:
                                usuario = usuarios.get(username_mapeado)
                                if usuario:
                                    usuario_encontrado = True
                                    break
                                else:
                                    print(f"Usuário mapeado {username_mapeado} não encontrado para tag {tag} no processo {numero_processo}")
                            else:
                                print(f"Tag não mapeada: {tag} para processo {numero_processo}")
                        if not usuario_encontrado:
                            usuario = None
                            print(f"Processo {numero_processo} sem usuário mapeado, nenhum usuário atribuído.")
                    else:
                        print(f"Processo {numero_processo} sem tags válidas ou tagsProcessoList vazio, usando usuário padrão {usuario_default}")
                        usuario = usuarios.get(usuario_default, User.objects.filter(username=usuario_default).first())

                    # Processar data de chegada
                    antigo = None
                    if pd.notna(row.get('dataChegada')) and row['dataChegada'].strip():
                        try:
                            data_str = str(row['dataChegada']).strip()
                            antigo = datetime.strptime(data_str, '%d/%m/%Y')
                            antigo = antigo.replace(hour=0, minute=0, second=0, microsecond=0)
                            antigo = timezone.make_aware(antigo)
                        except ValueError as e:
                            print(f"Erro ao converter data para {numero_processo}: {str(e)}")
                            messages.warning(request, f"Data inválida para processo {numero_processo}, ignorando data.")

                    # Processar prioridade
                    prioridade_str = str(row.get('prioridade', 'false')).strip().lower()
                    prioridade_urgente = prioridade_str == 'true'

                    # Garantir que processos já existentes NÃO percam o "tipo" original
                    campos_atualizaveis = {
                        'tema': tema,
                        'usuario': usuario,
                        'antigo': antigo,
                        'prioridade_urgente': prioridade_urgente,
                        'dt_atualizacao': timezone.now(),
                        'despacho': despacho,
                        'tags_materia': tags_materia,
                    }

                    processo, created = Processo.objects.get_or_create(
                        numero_processo=numero_processo,
                        especie=especie,
                        concluido=False,
                        defaults={
                            **campos_atualizaveis,
                            'dt_criacao': timezone.now(),
                            'data_dist': timezone.now(),
                            'tipo': tipo_processo, # O tipo só é definido na Criação!
                        }
                    )

                    if not created:
                        # Se já existe, atualizamos apenas os campos permitidos (sem mexer no tipo)
                        for campo, valor in campos_atualizaveis.items():
                            setattr(processo, campo, valor)
                        processo.save()

                    
                    if usuario:
                        # [CORREÇÃO]: Atualiza o responsável APENAS se o andamento 
                        # estiver na fase de "Elaboração".
                        # O termo fase__fase acessa o campo 'fase' do modelo Fase relacionado.
                        ProcessoAndamento.objects.filter(
                            processo=processo, 
                            fase__fase="Elaboração" 
                        ).update(usuario=usuario)
                    
                        processos_inseridos += 1
                        
                except Exception as e:
                    print(f"Erro ao importar processo {row.get('numeroProcesso', 'Desconhecido')}: {str(e)}")
                    # ... resto do código ...
                    

                except Exception as e:  # Linha 236 - Garantindo indentação correta
                    print(f"Erro ao importar processo {row.get('numeroProcesso', 'Desconhecido')}: {str(e)}")
                    messages.warning(request, f"Erro ao importar processo {row.get('numeroProcesso', 'Desconhecido')}: {str(e)}")
                    continue

            messages.success(
                request,
                f"Importação concluída: {processos_inseridos} processos inseridos, {processos_ignorados} processos ignorados."
            )
            return HttpResponseRedirect(request.path)

        except pd.errors.ParserError as e:
            messages.error(request, f"Erro ao ler o arquivo CSV: {str(e)}")
            return render(request, 'importar_processos.html')
        except DatabaseError as e:
            messages.error(request, f"Erro ao salvar no banco de dados: {str(e)}")
            return render(request, 'importar_processos.html')
        except Exception as e:
            messages.error(request, f"Erro ao processar o arquivo: {str(e)}")
            return render(request, 'importar_processos.html')

    return render(request, 'importar_processos.html')

locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')


@login_required
def definir_tema(request, pk):
    """Processa o formulário do modal e define/atualiza o Tema do Processo."""
    processo = get_object_or_404(Processo, pk=pk)
    
    if request.method == 'POST':
        tema_id = request.POST.get('tema_id')
        if tema_id:
            tema = get_object_or_404(Tema, pk=tema_id)
            processo.tema = tema
            processo.save()
            # Verificar se a requisição é AJAX
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Tema atualizado com sucesso!',
                    'tema_nome': tema.nome,
                })
            else:
                messages.success(request, "Tema atualizado com sucesso!")
                return redirect('processo_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Selecione um tema válido.',
                }, status=400)
            else:
                messages.error(request, "Selecione um tema válido.")
                return redirect('processo_list')
    
    return redirect('processo_list')

class ProcessoListView(LoginRequiredMixin, ListView):
    model = Processo
    template_name = 'processo_list.html'
    context_object_name = 'processos'
    paginate_by = 20

    def get_queryset(self):
    # Adicionamos o prefetch_related('comentarios') para performance
        queryset = Processo.objects.all().prefetch_related('comentarios')

    def get_queryset(self):
        """
        Filtra os processos com base nos parâmetros fornecidos na URL.
        Garante que processos na fase "Processo Concluído" apareçam quando o filtro de status for "Concluído".
        """
        queryset = Processo.objects.all()

        # 🔹 Captura filtros da URL
        despacho = self.request.GET.get('despacho')
        prioridade = self.request.GET.get('prioridade')
        status = self.request.GET.get('status', "pendente").strip().lower()
        fase_atual = self.request.GET.get('fase_atual')
        camara = self.request.GET.get('camara')
        tipo = self.request.GET.get('tipo')
        especie = self.request.GET.get('especie')
        numero_processo = self.request.GET.get('numero_processo')
        meus_processos = self.request.GET.get('meus_processos', None)
        user_id = self.request.GET.get('user_id')
        tema = self.request.GET.get('tema')

        # 🔹 Obtém a última fase do processo (incluindo todas as fases, independentemente do status)
        latest_fase = Subquery(
            ProcessoAndamento.objects.filter(
                processo=OuterRef('id')
            ).order_by('-dt_criacao').values('fase__fase')[:1]
        )

        # 🔹 Anota a fase atual (sem excluir "Processo Concluído")
        queryset = queryset.annotate(fase_atual=latest_fase)

        # 🔹 Filtra pelo status (pendente ou concluído)
        if status == "concluído":
            queryset = queryset.filter(concluido=True)
        else:
            queryset = queryset.filter(concluido=False)

        # 🔹 Aplica filtros adicionais
        if numero_processo:
            numero_processo = numero_processo.strip()[:10]
        filtros = {
            'fase_atual': fase_atual,
            'camara__camara': camara,
            'tipo__tipo': tipo,
            'especie__especie': especie,
            'numero_processo__icontains': numero_processo,
        }
        if tema:
            filtros['tema__nome'] = tema

        for campo, valor in filtros.items():
            if valor:
                queryset = queryset.filter(**{campo: valor})

        # 🔹 Filtrar processos apenas do usuário logado (se opção estiver ativada)
        if meus_processos == 'on':
            queryset = queryset.filter(usuario=self.request.user)

        # 🔹 Filtrar por usuário específico
        if user_id:
            queryset = queryset.filter(usuario__id=user_id)

        # 🔹 Filtragem por datas (se aplicável)
        data_filtros = {
            'data_dist': 'data_dist',
            'data_prazo': 'dt_prazo',
            'data_julgamento': 'dt_julgamento',
        }

        for param, field in data_filtros.items():
            valor = self.request.GET.get(param)
            if valor:
                queryset = queryset.filter(**{f"{field}__date": datetime.strptime(valor, '%Y-%m-%d').date()})

        # 🔹 Aplicar os filtros de despacho e prioridade
        if despacho == 'sim':
            queryset = queryset.filter(despacho=True)
        if prioridade == 'sim':
            queryset = queryset.filter(prioridade_urgente=True)

        # 🔹 Captura o parâmetro de ordenação da URL, com padrão "mais_recente"
        order_by = self.request.GET.get("ordenar", "mais_recente")

        # 🔹 Aplica ordenação segura
        ordering_dict = {
            "mais_recente": "-data_dist",
            "mais_antigo": "data_dist",
        }

        if order_by in ordering_dict:
            queryset = queryset.order_by(ordering_dict[order_by])

        # 🔹 Ordenação por 'dias_no_gabinete' (realiza sorting manual)
        elif order_by in ["dias_gabinete_recente", "dias_gabinete_antigo"]:
            queryset_list = list(queryset)
            reverse = order_by == "dias_gabinete_recente"
            queryset_list.sort(key=lambda p: p.dias_no_gabinete() or 0, reverse=reverse)
            return queryset_list

        return queryset

    def get_context_data(self, **kwargs):
        """
        Adiciona dados adicionais ao contexto do template.
        """
        context = super().get_context_data(**kwargs)

        # Captura o status da requisição ou define "Pendente" como padrão
        status_atual = self.request.GET.get('status', 'pendente')

        # Opções de ordenação
        context["ordenacao_opcoes"] = [
            {"valor": "mais_recente", "label": "Mais Recente"},
            {"valor": "mais_antigo", "label": "Mais Antigo"},
            {"valor": "dias_gabinete_recente", "label": "Mais Tempo no Gabinete"},
            {"valor": "dias_gabinete_antigo", "label": "Menos Tempo no Gabinete"},
        ]

        # Usuários ordenados
        context['users'] = User.objects.filter(
            id__in=Processo.objects.values_list('usuario_id', flat=True)
        ).select_related('profile').order_by('first_name', 'last_name')

        # Adiciona filtros para os campos relacionados
        context['statuses'] = ["Em andamento", "Não iniciado", "Concluído", "Pendente"]
        context['status_selecionado'] = status_atual
        context['fases'] = Fase.objects.exclude(fase="Concluído").values_list('fase', flat=True)
        context['camaras'] = Camara.objects.all()
        context['tipos'] = Tipo.objects.all()
        context['especies'] = Especie.objects.all()
        context['temas'] = Tema.objects.all()

        # Captura o tema selecionado para manter no dropdown
        context['tema_selecionado'] = self.request.GET.get('tema', '')

        # 🔹 Adiciona tarefas do dia do usuário
        tarefas = TarefaDoDia.objects.filter(usuario=self.request.user)
        context['tarefas_do_dia'] = tarefas
        context['tarefas_do_dia_ids'] = list(tarefas.values_list('processo__id', flat=True))
        context['despacho_selecionado'] = self.request.GET.get('despacho', 'nao')
        context['prioridade_selecionada'] = self.request.GET.get('prioridade', 'nao')

        # 🔹 Adiciona métricas de processos por usuário
        metrics_data = get_advanced_metrics()
        context["assessor_process_data"] = {
            item["id"]: item for item in metrics_data["assessor_process_data"]
        }

        return context


class ProcessoCreateView(CreateView):
    model = Processo
    form_class = ProcessoForm
    template_name = 'processo_form.html'  # sem extends, apenas o corpo do form
    success_url = reverse_lazy('processo_list')

    def form_invalid(self, form):
        """
        Se o formulário for inválido (erro de validação),
        retorne novamente o HTML do form (template) para exibir os erros dentro do modal.
        """
        # Basta fazer um render normal do template
        return self.render_to_response(self.get_context_data(form=form))

    def form_valid(self, form):
        """
        Se o formulário for válido, salve e retorne algo que indique 'sucesso' para o Ajax.
        """
        super().form_valid(form)
        # Aqui podemos retornar, por exemplo, apenas um status 204 (no content)
        return HttpResponse('', status=204)
        # ou um JsonResponse({'msg': 'criado com sucesso!'}, status=200)


class ProcessoDetailView(LoginRequiredMixin, DetailView):
    model = Processo
    template_name = 'processo_detail.html'
    context_object_name = 'processo'


class ProcessoUpdateView(LoginRequiredMixin, UpdateView):
    model = Processo
    form_class = ProcessoForm
    template_name = 'processo_form_update.html'
    success_url = reverse_lazy('processo_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Adiciona o usuário logado
        return kwargs


class ProcessoDeleteView(LoginRequiredMixin, DeleteView):
    model = Processo
    template_name = 'processo_confirm_delete.html'
    success_url = reverse_lazy('processo_list')


class AndamentoListView(LoginRequiredMixin, ListView):
    template_name = 'andamento_list.html'
    context_object_name = 'andamentos'

    def get_queryset(self):
        """
        Retorna os andamentos relacionados ao processo atual, com validação de ID.
        """
        processo_id = self.request.GET.get('processo')
        if not processo_id or not processo_id.isdigit():
            raise Http404("Processo inválido ou não encontrado.")
        
        from django.db.models import F
        return ProcessoAndamento.objects.filter(processo_id=processo_id).select_related('fase', 'usuario', 'status').order_by(
            F('dt_conclusao').asc(nulls_last=True),
            'dt_criacao'
        )

    def get_context_data(self, **kwargs):
        """
        Adiciona informações adicionais ao contexto, excluindo a fase 'Processo Concluído'.
        """
        context = super().get_context_data(**kwargs)

        processo_id = self.request.GET.get('processo')
        if not processo_id or not processo_id.isdigit():
            raise Http404("Processo inválido ou não encontrado.")

        # Busca o processo
        processo = get_object_or_404(Processo, pk=processo_id)

        # Exclui a fase "Processo Concluído"
        fases = Fase.objects.exclude(fase="Processo Concluído")

        # Organiza os andamentos por fase
        andamentos_por_fase = [
            {
                'fase': fase,
                'nao_iniciado_em_andamento': self.get_queryset().filter(fase=fase).exclude(status__status="Concluído"),
                'concluidos': self.get_queryset().filter(fase=fase, status__status="Concluído"),
            }
            for fase in fases
        ]

        # Criar o formulário com os dados do processo
        form = ProcessoForm(instance=processo)

        # Carregar opções corretamente para os dropdowns
        form.fields["tipo"].queryset = Tipo.objects.all()
        form.fields["resultado"].queryset = Resultado.objects.all()

        # DEBUG: Verifique se os valores estão sendo carregados corretamente
        print(f"Tipo Selecionado: {processo.tipo}, Resultado Selecionado: {processo.resultado}")

        ultimo_andamento = processo.andamentos.filter(
            status__status__in=["Não iniciado", "Em andamento"]
        ).order_by('dt_criacao').last()
        fase_atual_str = ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem Fase"

        # Atualiza o contexto com os dados do processo, andamentos e formulário
        context.update({
            'processo': processo,
            'andamentos_por_fase': andamentos_por_fase,
            'form': form,
            'fase_atual_str': fase_atual_str,
        })
        return context

class ProcessoPartialUpdateView(View):
    def post(self, request, processo_id):
        processo = get_object_or_404(Processo, id=processo_id)

        # Obtém os dados do POST e remove espaços em branco
        tipo_id = request.POST.get("tipo", "").strip()
        resultado_id = request.POST.get("resultado", "").strip()  # opcional
        despacho_val = request.POST.get("despacho")  # vem como "on" se marcado

        # Valida se o campo 'tipo' foi preenchido
        if not tipo_id:
            return JsonResponse(
                {'error': 'Por favor, selecione o Tipo antes de salvar.'},
                status=400
            )

        # Converte o id do tipo para instância do modelo
        tipo = get_object_or_404(Tipo, id=tipo_id)

        # Converte o id do resultado, se enviado; caso contrário, mantém como None
        if resultado_id:
            resultado = get_object_or_404(Resultado, id=resultado_id)
        else:
            resultado = None

        # Atualiza os valores no objeto processo
        processo.tipo = tipo
        processo.resultado = resultado
        processo.despacho = despacho_val == "on"  # Checkbox: True se marcado, False caso contrário
        processo.save()

        print(f"✅ Processo atualizado - Tipo: {processo.tipo}, Resultado: {processo.resultado}, Despacho: {processo.despacho}")

        # Redireciona para a página desejada com o objeto atualizado
        return redirect(f'/andamentos/?processo={processo.id}')

    

class ProcessoUpdateView(LoginRequiredMixin, UpdateView):
    model = Processo
    form_class = ProcessoForm
    template_name = 'processo_form_update.html'
    success_url = reverse_lazy('processo_list')

    # Passa o usuário logado para o formulário
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Adiciona o usuário logado
        return kwargs


from django.http import Http404


class AndamentoCreateView(LoginRequiredMixin,CreateView):
    model = ProcessoAndamento
    form_class = AndamentoForm
    template_name = 'andamento_form.html'
    success_url = reverse_lazy('andamento_list')


class AndamentoUpdateView(LoginRequiredMixin, UpdateView):
    model = ProcessoAndamento
    form_class = AndamentoForm
    template_name = 'andamento_form_update.html'

    def get_success_url(self):
        # Obtém o parâmetro 'origem' do formulário
        origem = self.request.POST.get('origem', 'andamento_list')
        # Obtém o processo associado ao andamento
        processo_id = self.object.processo.id

        if origem == 'home':
            base_url = reverse('home')
        else:
            base_url = f"{reverse('andamento_list')}?processo={processo_id}"

        # Se houver a query string dos filtros, adicione-a
        next_query = self.request.POST.get('next')
        if next_query:
            if '?' in base_url:
                return f"{base_url}&{next_query}"
            else:
                return f"{base_url}?{next_query}"
        return base_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["andamento"] = self.object
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Verifica se a ação é para iniciar o andamento
        if 'iniciar_andamento' in request.POST:
            if not self.object.dt_inicio:  # Evita reiniciar um andamento já iniciado
                self.object.dt_inicio = now()
                self.object.status = Status.objects.get(status="Em andamento")
                self.object.save()
            return redirect(self.get_success_url())
        
        # Verifica se a ação é para enviar para outra fase
        if 'enviar_para_fase' in request.POST:
            nova_fase = request.POST.get('nova_fase')
            if nova_fase:
                self.object.concluir_andamento()
                nova_fase_obj = Fase.objects.get(fase=nova_fase)
                ProcessoAndamento.objects.create(
                    processo=self.object.processo,
                    andamento=f"Movido para {nova_fase}",
                    fase=nova_fase_obj,
                    usuario=request.user,
                    status=Status.objects.get(status="Não iniciado")
                )
            return redirect(self.get_success_url())
        
        # Processa o formulário padrão (atualização via UpdateView)
        response = super().post(request, *args, **kwargs)
        return response


class AndamentoDeleteView(LoginRequiredMixin,DeleteView):
    model = ProcessoAndamento
    template_name = "andamento_confirm_delete.html"

    def get_success_url(self):
        """
        Redireciona para a lista de andamentos do processo atual após a exclusão.
        """
        processo_id = self.object.processo.id  # Obtém o ID do processo relacionado
        return reverse('andamento_list') + f'?processo={processo_id}'



class AndamentoIniciarView(LoginRequiredMixin,UpdateView):
    def post(self, request, pk, *args, **kwargs):
        andamento = get_object_or_404(ProcessoAndamento, pk=pk)
        if not andamento.dt_inicio:  # Verifica se o andamento ainda não foi iniciado
            andamento.dt_inicio = now()
            andamento.status = Status.objects.get(status="Em andamento")  # Atualiza o status
            andamento.save()
        return redirect(reverse('andamento_update', kwargs={'pk': pk}))


class AndamentoEnviarParaFaseView(LoginRequiredMixin, UpdateView):
    def post(self, request, pk, *args, **kwargs):
        andamento = get_object_or_404(ProcessoAndamento, pk=pk)

        # Verifica se o usuário logado é o atribuído ao andamento
        if request.user != andamento.usuario:
            return HttpResponseForbidden("Você não tem permissão para movimentar este andamento.")

        nova_fase = request.POST.get('nova_fase')
        origem = request.POST.get('origem', 'home')  # Padrão para 'home' se não fornecido

        # Finaliza o andamento atual
        andamento.dt_conclusao = now()
        andamento.status = Status.objects.get(status="Concluído")
        andamento.save()

        # Identifica o responsável baseado na fase
        usuario_responsavel = andamento.processo.usuario  # Usuário padrão (do processo)
        if nova_fase == "Revisão":
            usuario_responsavel = UserProfile.objects.filter(funcao="revisor(a)").first().user
        elif nova_fase == "Revisão Des":
            usuario_responsavel = UserProfile.objects.filter(funcao="Desembargador").first().user

        # Cria o novo andamento
        nova_fase_obj = get_object_or_404(Fase, fase=nova_fase)
        ProcessoAndamento.objects.create(
            processo=andamento.processo,
            andamento=f"Movido para {nova_fase}",
            fase=nova_fase_obj,
            usuario=usuario_responsavel,
            status=Status.objects.get(status="Não iniciado"),
            link_doc=andamento.link_doc  # Copia o link do documento
        )

        # Redireciona para a página de origem
        if origem == "andamento_list":
            return redirect(reverse('andamento_list') + f"?processo={andamento.processo.pk}")
        else:  # Inclui 'home' e qualquer outro caso
            return redirect('home')



class AndamentoConcluirProcessoView(LoginRequiredMixin, UpdateView):
    def post(self, request, pk, *args, **kwargs):
        andamento = get_object_or_404(ProcessoAndamento, pk=pk)
        
        # Verificar se o processo está na fase "L. PJE"
        

        # Finaliza o andamento atual
        andamento.dt_conclusao = now()
        andamento.status = Status.objects.get(status="Concluído")
        andamento.save()

        # Marca o processo como concluído e define a data de conclusão
        processo = andamento.processo
        processo.concluido = True
        processo.dt_atualizacao = now()
        processo.dt_conclusao = now()
        processo.save()

        # **Corrigir a fase do andamento "Processo concluído"**
        fase_concluido, _ = Fase.objects.get_or_create(fase="Processo Concluído")

        # Criar um novo andamento indicando que o processo foi concluído na fase correta
        novo_andamento = ProcessoAndamento.objects.create(
            processo=processo,
            andamento="Processo concluído",
            fase=fase_concluido, 
            usuario=request.user,
            status=Status.objects.get(status="Concluído"),
            dt_inicio=now(),
            dt_conclusao=now(),
            link_doc=andamento.link_doc  # Copiar o link do documento
        )

        # Remover o processo do "Meu Dia" (deletar o registro em TarefaDoDia)
        TarefaDoDia.objects.filter(usuario=request.user, processo=processo).delete()

        # Obtém o parâmetro 'origem' do formulário (pode ser 'home' ou 'andamento_list')
        origem = request.POST.get('origem', 'andamento_list')  # Default para 'andamento_list' se não especificado

        # Define a URL de redirecionamento com base na origem
        if origem == 'home':
            url = reverse('home')
        else:
            url = reverse('andamento_list') + f"?processo={andamento.processo.pk}"

        return redirect(url)


@login_required
@require_POST
def concluir_em_lote(request):
    """
    Conclui em lote todos os processos do usuário logado que estão na fase 'L. PJE'.
    Reutiliza a mesma lógica de AndamentoConcluirProcessoView.
    """
    from django.utils.timezone import now as tz_now

    # Busca todos os andamentos ativos na fase L. PJE do usuário logado
    andamentos_lpje = ProcessoAndamento.objects.filter(
        usuario=request.user,
        fase__fase='L. PJE',
        status__status__in=['Não iniciado', 'Em andamento'],
        processo__concluido=False,
    ).select_related('processo', 'fase', 'status')

    status_concluido = Status.objects.get(status='Concluído')
    fase_concluido, _ = Fase.objects.get_or_create(fase='Processo Concluído')

    total = 0
    for andamento in andamentos_lpje:
        # Finaliza o andamento atual
        andamento.dt_conclusao = tz_now()
        andamento.status = status_concluido
        andamento.save()

        # Marca o processo como concluído
        processo = andamento.processo
        processo.concluido = True
        processo.dt_atualizacao = tz_now()
        processo.dt_conclusao = tz_now()
        processo.save()

        # Cria andamento de "Processo Concluído"
        ProcessoAndamento.objects.create(
            processo=processo,
            andamento='Processo concluído',
            fase=fase_concluido,
            usuario=request.user,
            status=status_concluido,
            dt_inicio=tz_now(),
            dt_conclusao=tz_now(),
            link_doc=andamento.link_doc,
        )

        # Remove do "Meu Dia"
        TarefaDoDia.objects.filter(usuario=request.user, processo=processo).delete()

        total += 1

    if total > 0:
        messages.success(request, f'{total} processo{"s" if total > 1 else ""} concluído{"s" if total > 1 else ""} com sucesso!')
    else:
        messages.warning(request, 'Nenhum processo em L.PJE encontrado para concluir.')

    return redirect('home')


def process_metrics_view(request):

    # Obtém os filtros da URL
    assessor = request.GET.get('assessor')
    mes_distribuicao = request.GET.get('mes_distribuicao')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    numero_processo = request.GET.get('numero_processo')

    # Converte strings de data para objetos datetime (se fornecidos)
    if data_inicio:
        data_inicio = make_aware(datetime.strptime(data_inicio, "%Y-%m-%d"))
    if data_fim:
        data_fim = make_aware(datetime.strptime(data_fim, "%Y-%m-%d"))

    # Obtém os dados de métricas
    metrics_data = get_advanced_metrics(
        assessor=assessor, 
        mes_distribuicao=mes_distribuicao, 
        data_inicio=data_inicio, 
        data_fim=data_fim,
        status=status,
        numero_processo=numero_processo
    )

    # Lista de meses disponíveis para filtro
    months = [(i, month_name[i]) for i in range(1, 13)]

    # Lista de assessores distintos
    assessores = Processo.objects.values('usuario__id', 'usuario__first_name', 'usuario__last_name').distinct()

    # Passa os dados ao template
    return render(request, 'metrics.html', {
        # Dados de gráficos com quantidade e porcentagem
        "species_data": metrics_data["species_data"],    
        "type_data": metrics_data["type_data"],          
        "camara_data": metrics_data["camara_data"],      
        "resultado_data": metrics_data["resultado_data"],
        "assessor_data": metrics_data["assessor_data"],  
        "total_processos": metrics_data["total_processos"],
        "total_concluidos": metrics_data["total_concluidos"],
        "total_pendentes": metrics_data["total_pendentes"],
        "andamento_data": metrics_data["andamento_data"],
        "porcentagem_concluidos": metrics_data["porcentagem_concluidos"],
        "porcentagem_pendentes": metrics_data["porcentagem_pendentes"],

        # Métricas adicionais
        "average_process_time": metrics_data["average_process_time"], 
        "andamento_durations": metrics_data["andamento_durations"],  
        "andamento_waiting_times": metrics_data["andamento_waiting_times"], 

        # **Novo**: Lista de processos detalhados
        "detalhes_processos": metrics_data["detalhes_processos"],

        # 🔹 **Corrigido**: Passando os dados dos assessores corretamente
        "assessor_process_data": metrics_data["assessor_process_data"],  

        "months": months,
        "assessores": assessores,
        "filtros": {
            "data_inicio": request.GET.get('data_inicio', ''),
            "data_fim": request.GET.get('data_fim', ''),
            "status": request.GET.get('status', ''),
            "numero_processo": request.GET.get('numero_processo', '')
        }
    })


@login_required
def adicionar_tarefa(request, processo_id):
    processo = get_object_or_404(Processo, id=processo_id)
    TarefaDoDia.objects.get_or_create(usuario=request.user, processo=processo)

    # Obtém o parâmetro 'origem' do formulário (pode ser 'home' ou 'processo_list')
    origem = request.POST.get('origem', 'processo_list')  # Default para 'processo_list' se não especificado

    # Captura os parâmetros da URL para manter os filtros
    query_params = request.GET.copy()
    query_params.pop('page', None)  # Remove paginação para evitar problemas

    # Define a URL de redirecionamento com base na origem
    if origem == 'home':
        url = reverse('home')
    else:
        url = reverse('processo_list')

    # Adiciona os parâmetros de query à URL, se existirem
    if query_params:
        return redirect(f"{url}?{query_params.urlencode()}")
    return redirect(url)

@login_required
def remover_tarefa(request, processo_id):
    processo = get_object_or_404(Processo, id=processo_id)
    TarefaDoDia.objects.filter(usuario=request.user, processo=processo).delete()

    # Obtém o parâmetro 'origem' do formulário (pode ser 'home' ou 'processo_list')
    origem = request.POST.get('origem', 'processo_list')  # Default para 'processo_list' se não especificado

    # Captura os parâmetros da URL para manter os filtros
    query_params = request.GET.copy()
    query_params.pop('page', None)  # Remove paginação para evitar problemas

    # Define a URL de redirecionamento com base na origem
    if origem == 'home':
        url = reverse('home')
    else:
        url = reverse('processo_list')

    # Adiciona os parâmetros de query à URL, se existirem
    if query_params:
        return redirect(f"{url}?{query_params.urlencode()}")
    return redirect(url)

from django.http import HttpResponse
from django.db.models.functions import Left
import openpyxl
from openpyxl.styles import Font, Alignment

def export_processos_xlsx(request):
    """Exporta os processos para um arquivo Excel (.xlsx) com base nos filtros aplicados na ProcessoListView"""
    
    # Instancia a ProcessoListView para reutilizar a lógica de filtros
    view = ProcessoListView()
    view.request = request  # Passa a requisição para a view
    processos = view.get_queryset()

    # 🔹 Obtém o link_doc do andamento com fase "L. PJE" (o mais recente, se houver múltiplos)
    l_pje_link_doc = Subquery(
        ProcessoAndamento.objects.filter(
            processo=OuterRef('id'),
            fase__fase="L. PJE"  # Filtra pela fase "L. PJE"
        ).order_by('-dt_criacao').values('link_doc')[:1]
    )

    # 🔹 Otimiza a query com select_related e anota o link_doc
    processos = processos.select_related('especie', 'usuario', 'resultado', 'tipo', 'camara', 'tema').annotate(
        l_pje_link_doc=l_pje_link_doc
    )

    # Criar um novo workbook (arquivo Excel)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"

    # Definir os cabeçalhos da planilha, incluindo as novas colunas
    headers = [
        "Número do Processo", "Data de Distribuição", "Espécie", "Resultado", "Tipo",
        "Câmara", "Usuário Responsável", "Data de Julgamento", "Prazo",
        "Data de Criação", "Última Atualização", "Concluído", "Data de Conclusão",
        "Tema", "Despacho", "Link do Documento"
    ]
    ws.append(headers)

    # Verifica se há processos; se não, adiciona uma mensagem
    if not processos:
        ws.append(["Nenhum processo encontrado com os filtros aplicados."])
    else:
        # Adicionar os dados à planilha
        for processo in processos:
            ws.append([
                processo.numero_processo,
                processo.data_dist.strftime('%d/%m/%Y %H:%M') if processo.data_dist else "",
                processo.especie.especie if processo.especie else "",
                processo.resultado.resultado if processo.resultado else "",
                processo.tipo.tipo if processo.tipo else "",
                processo.camara.camara if processo.camara else "",
                f"{processo.usuario.first_name} {processo.usuario.last_name}" if processo.usuario else "Sem responsável",
                processo.dt_julgamento.strftime('%d/%m/%Y %H:%M') if processo.dt_julgamento else "",
                processo.dt_prazo.strftime('%d/%m/%Y %H:%M') if processo.dt_prazo else "",
                processo.dt_criacao.strftime('%d/%m/%Y %H:%M'),
                processo.dt_atualizacao.strftime('%d/%m/%Y %H:%M'),
                "Sim" if processo.concluido else "Não",
                processo.dt_conclusao.strftime('%d/%m/%Y %H:%M') if processo.dt_conclusao else "",
                processo.tema.nome if processo.tema else "",  # Novo campo: Tema
                "Sim" if processo.despacho else "Não",        # Novo campo: Despacho
                processo.l_pje_link_doc or ""                 # Novo campo: Link do Documento (fase L. PJE)
            ])

    # Estilizar cabeçalhos
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Criar a resposta HTTP com o arquivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="processos.xlsx"'
    wb.save(response)

    return response

@login_required
def adicionar_comentario(request, processo_id):
    processo = get_object_or_404(Processo, id=processo_id)
    
    # ... (mantenha suas verificações de permissão aqui)

    if request.method == "POST":
        form = ComentarioProcessoForm(request.POST)
        if form.is_valid():
            comentario = form.save(commit=False)
            comentario.processo = processo
            comentario.usuario = request.user
            comentario.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                autor_e_revisor = hasattr(request.user, 'profile') and request.user.profile.funcao == "revisor(a)"
                return JsonResponse({
                    "status": "success",
                    "texto": comentario.texto,
                    "usuario": request.user.get_full_name() or request.user.username,
                    "data_criacao": comentario.data_criacao.strftime('%d/%m/%Y %H:%M'),
                    "photo_url": request.user.profile.photo.url if hasattr(request.user, 'profile') and request.user.profile.photo else None,
                    "funcao": request.user.profile.funcao
                })
    
    return redirect(f'/andamentos/?processo={processo_id}')

def preparar_processos_por_usuario(hoje):
    usuarios = User.objects.all()
    processos_abertos = Processo.objects.filter(concluido=False).select_related('especie', 'usuario', 'tipo')

    ultima_fase_subquery = ProcessoAndamento.objects.filter(
        processo=OuterRef('pk')
    ).order_by('-dt_criacao').values('fase__fase')[:1]

    processos_abertos = processos_abertos.annotate(fase_atual=Subquery(ultima_fase_subquery))

    processos_por_usuario = {}
    for usuario in usuarios:
        processos_usuario = processos_abertos.filter(usuario=usuario)
        processos_por_usuario[usuario.id] = [
            {
                'id': p.id, # Usando 'p' conforme o loop
                'numero_processo': p.numero_processo,
                'especie': p.especie.especie if p.especie else '—',
                'tipo': p.tipo.tipo if p.tipo else 'Não informado',
                'dias_no_gabinete': (hoje - p.antigo.date()).days if p.antigo else 0,
                'fase_atual': p.fase_atual or '—'
            }
            for p in processos_usuario
        ]
    return processos_por_usuario


class ProcessosCreateListAPIView(generics.ListCreateAPIView):
    queryset = models.Processo.objects.all()
    serializer_class = serializers.ProcessoSerializer

@login_required
def excluir_comentario(request, comentario_id):
    comentario = get_object_or_404(ComentarioProcesso, id=comentario_id)
    processo_id = comentario.processo.id
    
    # Validação de segurança: apenas o perfil com função "Desembargador" pode apagar
    if hasattr(request.user, 'profile') and request.user.profile.funcao == "Desembargador":
        comentario.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Comentário removido.'})
        messages.success(request, "Comentário removido com sucesso.")
    else:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Acesso negado.'}, status=403)
        messages.error(request, "Apenas o Desembargador pode excluir comentários.")
    
    return redirect(f'/andamentos/?processo={processo_id}')

def get_queryset(self):
        queryset = models.Processo.objects.all()
        request = self.request

        numero = request.GET.get('numero_processo')
        concluido = request.GET.get('concluido')

        if numero:
            queryset = queryset.filter(numero_processo=numero)

        if concluido is not None:
            concluido_bool = concluido.lower() == 'true'
            queryset = queryset.filter(concluido=concluido_bool)

        return queryset


User = get_user_model()

@login_required
@require_http_methods(["GET"])
def listar_metas_semanal(request):
    """
    View melhorada para listagem de metas semanais com:
    - Dashboard de estatísticas
    - Filtros avançados
    - Paginação
    - Status calculados
    - Performance otimizada
    """
    # 1) Define a data atual para comparações
    agora = timezone.localtime()
    hoje = timezone.localdate()

    # 2) Obtém todas as metas disponíveis com permissões
    if request.user.has_perm('app_name.view_all_metas'):
        todas_metas = MetaSemanal.objects.all()
    else:
        todas_metas = MetaSemanal.objects.filter(usuario=request.user)

    # 3) Cria lista de períodos disponíveis (otimizada)
    periodos_disponiveis = []
    semanas_unicas = todas_metas.values('semana_inicio', 'semana_fim').distinct().order_by('semana_inicio')
    
    for semana in semanas_unicas:
        inicio = semana['semana_inicio']
        fim = semana['semana_fim']
        periodos_disponiveis.append({
            'inicio': inicio,
            'fim': fim,
            'label': f"{inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}",
            'value': f"{inicio.strftime('%Y-%m-%d')}|{fim.strftime('%Y-%m-%d')}"
        })

    # 4) Aplicar filtros
    metas_filtradas = aplicar_filtros(request, todas_metas, hoje)
    
    # 5) Calcular estatísticas para o dashboard
    estatisticas = calcular_estatisticas(metas_filtradas, hoje)
    
    # 6) Preparar dados de processos por usuário
    processos_por_usuario = preparar_processos_por_usuario(hoje)
    
    # 7) Calcular progresso e métricas para cada meta
    metas_com_dados = calcular_dados_metas(metas_filtradas, hoje)
    
    # 8) Aplicar paginação
    paginator = Paginator(metas_com_dados, 12)  # 12 metas por página
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # 9) Preparar lista de usuários para filtro
    usuarios_disponiveis = User.objects.filter(
        id__in=todas_metas.values_list('usuario_id', flat=True)
    ).order_by('first_name', 'last_name')

    # 10) Contexto para o template
    context = {
        'metas': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'processos_por_usuario': processos_por_usuario,
        'periodos': periodos_disponiveis,
        'usuarios': usuarios_disponiveis,
        'estatisticas': estatisticas,
        
        # Filtros aplicados (para manter estado)
        'periodo_selecionado': request.GET.get('periodo', ''),
        'usuario_selecionado': request.GET.get('usuario', ''),
        'status_selecionado': request.GET.get('status', ''),
        'busca_termo': request.GET.get('busca', ''),
    }

    return render(request, 'listar_metas_semanal.html', context)


def aplicar_filtros(request, queryset, hoje):
    """
    Aplica filtros avançados na queryset de metas
    """
    # Filtro por período
    periodo_selecionado = request.GET.get('periodo')
    if periodo_selecionado:
        try:
            inicio_str, fim_str = periodo_selecionado.split('|')
            inicio_semana = datetime.strptime(inicio_str, '%Y-%m-%d').date()
            fim_semana = datetime.strptime(fim_str, '%Y-%m-%d').date()
            queryset = queryset.filter(semana_inicio=inicio_semana, semana_fim=fim_semana)
        except (ValueError, TypeError):
            pass  # Ignora filtro malformado
    
    # Filtro por usuário
    usuario_selecionado = request.GET.get('usuario')
    if usuario_selecionado:
        try:
            user_id = int(usuario_selecionado)
            queryset = queryset.filter(usuario_id=user_id)
        except (ValueError, TypeError):
            pass
    
    # Filtro por busca (nome do usuário)
    busca_termo = request.GET.get('busca', '').strip()
    if busca_termo:
        queryset = queryset.filter(
            Q(usuario__first_name__icontains=busca_termo) |
            Q(usuario__last_name__icontains=busca_termo)
        )
    
    # Aplicar select_related e prefetch_related para otimização
    queryset = queryset.select_related('usuario', 'usuario__profile').prefetch_related(
        Prefetch('processos', Processo.objects.select_related('especie'))
    )
    
    return queryset


def calcular_estatisticas(metas, hoje):
    """
    Calcula estatísticas para o dashboard
    """
    if not metas.exists():
        return {
            'total_usuarios': 0,
            'metas_atingidas': 0,
            'metas_andamento': 0,
            'media_progresso': 0
        }
    
    # Calcular progresso para cada meta (versão simplificada)
    metas_com_progresso = []
    for meta in metas:
        progresso = calcular_progresso_meta(meta, hoje)
        metas_com_progresso.append({
            'meta': meta,
            'progresso': progresso,
            'status': determinar_status_meta(meta, hoje, progresso)
        })
    
    # Calcular estatísticas
    total_usuarios = len(set(m['meta'].usuario_id for m in metas_com_progresso))
    metas_atingidas = sum(1 for m in metas_com_progresso if m['progresso'] >= 100)
    metas_andamento = sum(1 for m in metas_com_progresso if 0 < m['progresso'] < 100)
    media_progresso = sum(m['progresso'] for m in metas_com_progresso) / len(metas_com_progresso)
    
    return {
        'total_usuarios': total_usuarios,
        'metas_atingidas': metas_atingidas,
        'metas_andamento': metas_andamento,
        'media_progresso': round(media_progresso, 1)
    }


def preparar_processos_por_usuario(hoje):
    """
    Prepara dicionário de processos abertos por usuário
    """
    usuarios = User.objects.all()
    processos_abertos = Processo.objects.filter(concluido=False).select_related('especie', 'usuario', 'tipo')

    # Subquery para fase atual
    ultima_fase_subquery = ProcessoAndamento.objects.filter(
        processo=OuterRef('pk')
    ).order_by('-dt_criacao').values('fase__fase')[:1]

    processos_abertos = processos_abertos.annotate(fase_atual=Subquery(ultima_fase_subquery))

    processos_por_usuario = {}
    for usuario in usuarios:
        processos_usuario = processos_abertos.filter(usuario=usuario)
        processos_por_usuario[usuario.id] = [
            {
                'id': p.id,
                'numero_processo': p.numero_processo,
                'especie': p.especie.especie if p.especie else '—',
                'tipo': p.tipo.tipo if p.tipo else 'Não informado', # Adicionado aqui
                'dias_no_gabinete': (hoje - p.antigo.date()).days if p.antigo else 0,
                'fase_atual': p.fase_atual or '—'
            }
            for p in processos_usuario
        ]
    
    return processos_por_usuario


def calcular_dados_metas(metas, hoje):
    """
    Calcula progresso, status e métricas para cada meta
    """
    metas_processadas = []
    
    for meta in metas:
        # Calcular progresso
        progresso = calcular_progresso_meta(meta, hoje)
        
        # Determinar status
        status = determinar_status_meta(meta, hoje, progresso)
        
        # Calcular faixas de dias
        faixas = calcular_faixas_dias(meta, hoje)
        
        # Adicionar dados calculados à meta
        meta.progresso = progresso
        meta.status_calculado = status
        meta.range_le30 = faixas['le30']
        meta.range_31_40 = faixas['31_40']
        meta.range_41_50 = faixas['41_50']
        meta.range_gt50 = faixas['gt50']
        
        # Calcular totais
        processos_meta = list(meta.processos.all())
        meta.total_processos = len(processos_meta)
        meta.processos_concluidos = calcular_processos_concluidos(meta, hoje)
        meta.processos_pendentes = meta.total_processos - meta.processos_concluidos
        
        # Determinar status da semana
        if meta.semana_fim < hoje:
            meta.status = 'passada'
        elif meta.semana_inicio > hoje:
            meta.status = 'futura'
        else:
            meta.status = 'atual'
        
        metas_processadas.append(meta)
    
    # Aplicar filtro por status se especificado
    status_filtro = None  # Você pode pegar do request se necessário
    if status_filtro:
        metas_processadas = [m for m in metas_processadas if m.status_calculado == status_filtro]
    
    return metas_processadas


def calcular_progresso_meta(meta, hoje):
    from django.db.models import Subquery, OuterRef
    FASES_EXATAS = ["Revisão", "Revisão Des", "L. PJE", "L.PJE", "Processo Concluído"]

    processos_meta_ids = list(meta.processos.values_list('id', flat=True))
    if not processos_meta_ids:
        return 0

    # 1. Base Histórica
    minutados_periodo = set(
        ProcessoAndamento.objects.filter(
            fase__fase__in=FASES_EXATAS,
            processo_id__in=processos_meta_ids,
            dt_criacao__date__range=(meta.semana_inicio, meta.semana_fim)
        ).values_list('processo_id', flat=True).distinct()
    )

    if meta.semana_fim >= hoje:
        # 2. Base Atual
        ultima_fase_sub = ProcessoAndamento.objects.filter(
            processo=OuterRef('pk')
        ).order_by('-dt_criacao').values('fase__fase')[:1]

        minutados_fase = set(
            Processo.objects.filter(
                id__in=processos_meta_ids
            ).annotate(
                fase_atual=Subquery(ultima_fase_sub)
            ).filter(
                fase_atual__in=FASES_EXATAS
            ).values_list('id', flat=True)
        )
        concluidas = len(minutados_periodo | minutados_fase)
    else:
        concluidas = len(minutados_periodo)

    progresso = round(min((concluidas / meta.meta_qtd * 100) if meta.meta_qtd else 0, 100), 1)
    return progresso


def determinar_status_meta(meta, hoje, progresso):
    """
    Determina o status calculado da meta
    """
    if meta.semana_fim < hoje:
        # Meta passada
        return 'atingida' if progresso >= 100 else 'atrasada'
    elif meta.semana_inicio > hoje:
        # Meta futura
        return 'futura'
    else:
        # Meta atual
        if progresso >= 100:
            return 'atingida'
        elif progresso >= 80:
            return 'quase_la'
        else:
            return 'em_andamento'


def calcular_faixas_dias(meta, hoje):
    """
    Calcula distribuição de processos por faixas de dias
    """
    processos_usuario = Processo.objects.filter(
        usuario=meta.usuario,
        concluido=False
    ).exclude(antigo__isnull=True)
    
    total_user = processos_usuario.count()
    if total_user == 0:
        return {'le30': 0, '31_40': 0, '41_50': 0, 'gt50': 0}
    
    dias = [(hoje - p.antigo.date()).days for p in processos_usuario]
    mais_30 = sum(1 for d in dias if d > 30)
    mais_40 = sum(1 for d in dias if d > 40)
    mais_50 = sum(1 for d in dias if d > 50)

    return {
        'le30': total_user - mais_30,
        '31_40': mais_30 - mais_40,
        '41_50': mais_40 - mais_50,
        'gt50': mais_50
    }


def calcular_processos_concluidos(meta, hoje):
    from django.db.models import Subquery, OuterRef
    FASES_EXATAS = ["Revisão", "Revisão Des", "L. PJE", "L.PJE", "Processo Concluído"]

    processos_meta_ids = list(meta.processos.values_list('id', flat=True))
    if not processos_meta_ids:
        return 0

    # 1. Base Histórica
    minutados_periodo = set(
        ProcessoAndamento.objects.filter(
            fase__fase__in=FASES_EXATAS,
            processo_id__in=processos_meta_ids,
            dt_criacao__date__range=(meta.semana_inicio, meta.semana_fim)
        ).values_list('processo_id', flat=True).distinct()
    )

    if meta.semana_fim >= hoje:
        # 2. Base Atual
        ultima_fase_sub = ProcessoAndamento.objects.filter(
            processo=OuterRef('pk')
        ).order_by('-dt_criacao').values('fase__fase')[:1]

        minutados_fase = set(
            Processo.objects.filter(
                id__in=processos_meta_ids
            ).annotate(
                fase_atual=Subquery(ultima_fase_sub)
            ).filter(
                fase_atual__in=FASES_EXATAS
            ).values_list('id', flat=True)
        )
        return len(minutados_periodo | minutados_fase)

    return len(minutados_periodo)

@login_required
@require_http_methods(["GET"])
def api_meta_detalhes(request):
    """
    API para retornar detalhes de uma meta específica
    """
    meta_id = request.GET.get('meta_id')
    if not meta_id:
        return JsonResponse({'error': 'meta_id é obrigatório'}, status=400)
    
    try:
        meta = get_object_or_404(MetaSemanal, id=meta_id)
        
        # Verificar permissões
        if not request.user.has_perm('app_name.view_all_metas') and meta.usuario != request.user:
            return JsonResponse({'error': 'Sem permissão'}, status=403)
        
        hoje = timezone.localdate()
        progresso = calcular_progresso_meta(meta, hoje)
        faixas = calcular_faixas_dias(meta, hoje)
        
        data = {
            'id': meta.id,
            'usuario': meta.usuario.get_full_name(),
            'periodo': f"{meta.semana_inicio.strftime('%d/%m/%Y')} a {meta.semana_fim.strftime('%d/%m/%Y')}",
            'meta_qtd': meta.meta_qtd,
            'progresso': progresso,
            'range_le30': faixas['le30'],
            'range_31_40': faixas['31_40'],
            'range_41_50': faixas['41_50'],
            'range_gt50': faixas['gt50'],
            'processos_total': meta.processos.count(),
            'processos_concluidos': calcular_processos_concluidos(meta, hoje),
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_meta_processos(request):
    """
    API para retornar processos de uma meta específica
    """
    meta_id = request.GET.get('meta_id')
    if not meta_id:
        return JsonResponse({'error': 'meta_id é obrigatório'}, status=400)
    
    try:
        meta = get_object_or_404(MetaSemanal, id=meta_id)
        
        # Verificar permissões
        if not request.user.has_perm('app_name.view_all_metas') and meta.usuario != request.user:
            return JsonResponse({'error': 'Sem permissão'}, status=403)
        
        hoje = timezone.localdate()
        
        # Buscar processos com fase atual
        processos = meta.processos.select_related('especie').annotate(
            fase_atual=Subquery(
                ProcessoAndamento.objects.filter(
                    processo=OuterRef('pk')
                ).order_by('-dt_criacao').values('fase__fase')[:1]
            )
        )
        
        FASES_EXATAS = ["Revisão", "Revisão Des", "L. PJE", "L.PJE", "Processo Concluído"]

        # Verificar quais estavam concluídos DENTRO do período da meta
        revisados_ids = set(
            ProcessoAndamento.objects.filter(
                fase__fase__in=FASES_EXATAS,
                processo__in=processos,
                dt_criacao__date__range=(meta.semana_inicio, meta.semana_fim)
            ).values_list('processo_id', flat=True).distinct()
        )
        
        is_meta_ativa = meta.semana_fim >= hoje
        processos_data = []
        concluidos_count = 0
        for processo in processos:
            dias_gabinete = (hoje - processo.antigo.date()).days if processo.antigo else 0
            
            is_concluido = processo.id in revisados_ids
            
            # Se a meta for da semana atual, considera a fase atual também
            if not is_concluido and is_meta_ativa:
                fase_raw = (processo.fase_atual or '')
                fase_limpa = fase_raw.strip()
                if fase_limpa in FASES_EXATAS:
                    is_concluido = True

            if is_concluido:
                concluidos_count += 1

            processos_data.append({
                'id': processo.id,
                'numero_processo': processo.numero_processo,
                'especie': processo.especie.especie if processo.especie else 'N/A',
                'fase_atual': processo.fase_atual or 'N/A',
                'dias_no_gabinete': dias_gabinete,
                'concluido': is_concluido
            })

        # Ordenar: concluídos primeiro, depois por dias no gabinete
        processos_data.sort(key=lambda x: (not x['concluido'], -x['dias_no_gabinete']))

        
        return JsonResponse({
            'processos': processos_data,
            'total': len(processos_data),
            'concluidos': concluidos_count
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def exportar_metas_relatorio(request):
    """
    Exporta relatório de metas em formato PDF
    """
    try:
        # Buscar metas com filtros aplicados
        if request.user.has_perm('app_name.view_all_metas'):
            metas = MetaSemanal.objects.all()
        else:
            metas = MetaSemanal.objects.filter(usuario=request.user)

        metas = aplicar_filtros(request, metas, timezone.localdate())
        metas_com_dados = calcular_dados_metas(metas, timezone.localdate())

        hoje = timezone.localdate()

        # Normalização de tipos com erros de digitação conhecidos
        TIPO_NORMALIZACAO = {
            'monocrátical': 'Monocrática',
            'monocratical': 'Monocrática',
            'urgentissimo': 'Urgentíssimo',
        }

        # Fases válidas para exibição — evita mistura de colunas
        FASES_VALIDAS_DISPLAY = {
            "Elaboração", "Revisão", "Revisão Des", "Correção",
            "Devolvido", "L. PJE", "PJE", "Processo Concluído", "Correção"
        }

        for meta in metas_com_dados:
            processos = meta.processos.select_related('especie', 'tipo').annotate(
                fase_atual=Subquery(
                    ProcessoAndamento.objects.filter(
                        processo=OuterRef('pk')
                    ).order_by('-dt_criacao').values('fase__fase')[:1]
                )
            )

            FASES_EXATAS = ["Revisão", "Revisão Des", "L. PJE", "L.PJE", "Processo Concluído"]

            # Minutados = tiveram andamento de conclusão DENTRO do período da meta
            revisados_ids = set(
                ProcessoAndamento.objects.filter(
                    fase__fase__in=FASES_EXATAS,
                    processo__in=processos,
                    dt_criacao__date__range=(meta.semana_inicio, meta.semana_fim)
                ).values_list('processo_id', flat=True).distinct()
            )

            is_meta_ativa = meta.semana_fim >= hoje
            processos_data = []

            for processo in processos:
                dias_gabinete = (hoje - processo.antigo.date()).days if processo.antigo else 0
                
                concluido = processo.id in revisados_ids
                fase_raw = (processo.fase_atual or '')
                
                if not concluido and is_meta_ativa:
                    fase_limpa = fase_raw.strip()
                    if fase_limpa in FASES_EXATAS:
                        concluido = True

                if dias_gabinete <= 30:   badge_class = "bg-success"
                elif dias_gabinete <= 40: badge_class = "bg-primary"
                elif dias_gabinete <= 50: badge_class = "bg-warning"
                else:                     badge_class = "bg-danger"

                # Sanitização do tipo
                tipo_nome_raw = (processo.tipo.tipo if processo.tipo else '') or ''
                tipo_nome = TIPO_NORMALIZACAO.get(tipo_nome_raw.lower().strip(), tipo_nome_raw)
                if not tipo_nome and getattr(processo, 'prioridade_urgente', False):
                    tipo_nome = 'Urgentíssimo'

                tipo_nome_lower = tipo_nome.lower()
                if 'urgentíssimo' in tipo_nome_lower or 'urgente' in tipo_nome_lower:
                    tipo_badge = 'bg-danger'
                elif 'prioridade' in tipo_nome_lower:
                    tipo_badge = 'bg-warning'
                elif 'monocrática' in tipo_nome_lower:
                    tipo_badge = 'bg-primary'
                elif tipo_nome:
                    tipo_badge = 'bg-secondary'
                else:
                    tipo_badge = ''

                # Sanitização da fase — evita mistura de colunas (fase_raw já definido acima)
                fase_exibir = fase_raw if fase_raw in FASES_VALIDAS_DISPLAY else (fase_raw or 'Não especificado')

                is_urgentissimo = 'urgentíssimo' in tipo_nome_lower or 'urgente' in tipo_nome_lower

                processos_data.append({
                    'numero_processo': processo.numero_processo,
                    'especie': processo.especie.especie if processo.especie else 'N/A',
                    'fase_atual': fase_exibir,
                    'dias_gabinete': dias_gabinete,
                    'badge_class': badge_class,
                    'concluido': concluido,
                    'status_texto': 'Minutado' if concluido else 'Pendente',
                    'tipo_nome': tipo_nome,
                    'tipo_badge': tipo_badge,
                    'is_urgentissimo': is_urgentissimo,
                })

            # Pendentes primeiro (urgentíssimos no topo) → concluídos ao final
            processos_data.sort(key=lambda x: (
                x['concluido'],            # False (pendente) vem antes de True (minutado)
                not x['is_urgentissimo'],  # Entre pendentes: urgentíssimos primeiro
                -x['dias_gabinete'],       # Depois: mais antigos primeiro
            ))

            meta.processos_data = processos_data
            meta.meta_qtd_exibir = meta.meta_qtd

        # ── DASHBOARD EXECUTIVO GLOBAL ─────────────────────────────────────
        total_processos_gab = sum(m.total_processos for m in metas_com_dados)
        total_minutados_gab = sum(m.processos_concluidos for m in metas_com_dados)
        total_pendentes_gab = sum(m.processos_pendentes for m in metas_com_dados)
        total_meta_qtd_gab  = sum(m.meta_qtd for m in metas_com_dados)

        taxa_cumprimento = round(
            (total_minutados_gab / total_meta_qtd_gab * 100), 1
        ) if total_meta_qtd_gab > 0 else 0.0

        urg_total = urg_minutados = 0
        for meta in metas_com_dados:
            for p in (meta.processos_data or []):
                if p['is_urgentissimo']:
                    urg_total += 1
                    if p['concluido']:
                        urg_minutados += 1

        metas_atingidas = sum(1 for m in metas_com_dados if m.progresso >= 100)
        metas_total = len(metas_com_dados)

        dashboard_global = {
            'total_processos': total_processos_gab,
            'total_minutados': total_minutados_gab,
            'total_pendentes': total_pendentes_gab,
            'taxa_cumprimento': taxa_cumprimento,
            'urgentissimos_total': urg_total,
            'urgentissimos_minutados': urg_minutados,
            'metas_atingidas': metas_atingidas,
            'metas_total': metas_total,
        }
        # ───────────────────────────────────────────────────────────────────

        import os
        from django.conf import settings
        logo_path = os.path.join(settings.MEDIA_ROOT, 'profile_photos', 'Logo.jpg')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.MEDIA_ROOT, 'banners', 'logo2.png')
            if not os.path.exists(logo_path):
                logo_path = None

        context = {
            'metas_com_dados': metas_com_dados,
            'dashboard_global': dashboard_global,
            'hoje': timezone.now(),
            'today': timezone.now(),
            'logo_path': logo_path,
        }

        from django.template.loader import get_template
        from xhtml2pdf import pisa
        from io import BytesIO

        template = get_template('metas_semanais_pdf.html')
        html = template.render(context)
        result = BytesIO()

        pisa_status = pisa.CreatePDF(html, dest=result, encoding='utf-8')

        if pisa_status.err:
            return HttpResponse('Erro ao gerar PDF', status=500)

        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"metas_semanais_{timezone.localdate().strftime('%Y-%m-%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        return JsonResponse({'error': f'Erro ao gerar relatório: {str(e)}'}, status=500)

@require_http_methods(["GET", "POST"])
@login_required
def editar_meta_semanal(request):
    if request.method == "GET":
        meta_id = request.GET.get('meta_id')
        meta = get_object_or_404(MetaSemanal, id=meta_id)

        # Verificar permissões
        if not request.user.has_perm('app_name.change_metasemanal') and request.user != meta.usuario:
            return JsonResponse({'success': False, 'message': 'Você não tem permissão para editar esta meta.'}, status=403)

        processos = list(meta.processos.all())

        # Subquery para fase atual
        ultima_fase_subquery = ProcessoAndamento.objects.filter(
            processo=OuterRef('pk')
        ).order_by('-dt_criacao').values('fase__fase')[:1]

        processos_com_fase = Processo.objects.filter(pk__in=[p.id for p in processos]).annotate(
            fase_atual=Subquery(ultima_fase_subquery)
        ).select_related('especie', 'tipo')

        data_processos = []
        for p in processos_com_fase:
            data_processos.append({
                'id': p.id,
                'numero_processo': p.numero_processo,
                'especie': p.especie.especie if p.especie else '—',
                'tipo': p.tipo.tipo if p.tipo else 'Não informado',
                'fase_atual': p.fase_atual or 'Não especificado',
                'dias_no_gabinete': (timezone.now().date() - p.antigo.date()).days if p.antigo else 0
            })

        data = {
            'id': meta.id,
            'usuario_id': meta.usuario.id,
            'meta_qtd': meta.meta_qtd,
            'processo_ids': [p.id for p in processos],
            'processos': data_processos
        }
        return JsonResponse(data)

    if request.method == "POST":
        meta_id = request.POST.get('meta_id')
        meta_qtd = request.POST.get('meta_qtd')
        processo_ids = request.POST.getlist('processo_ids[]')

        if not meta_id or not processo_ids:
            return JsonResponse({'success': False, 'message': 'Dados inválidos: meta e processos são obrigatórios.'}, status=400)

        try:
            meta_qtd = int(meta_qtd)
            if meta_qtd <= 0:
                return JsonResponse({'success': False, 'message': 'A meta deve ser maior que 0.'}, status=400)
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'message': 'A meta deve ser um número inteiro válido.'}, status=400)

        meta = get_object_or_404(MetaSemanal, id=meta_id)

        # Verificar permissões
        if not request.user.has_perm('app_name.change_metasemanal') and request.user != meta.usuario:
            return JsonResponse({'success': False, 'message': 'Você não tem permissão para editar esta meta.'}, status=403)

        # Inclui processos válidos (não concluídos e do usuário)
        processos_validos = Processo.objects.filter(id__in=processo_ids, usuario=meta.usuario)

        # Inclui também os concluídos que já estavam na meta
        processos_concluidos_meta = meta.processos.filter(id__in=processo_ids)

        # Combina os dois conjuntos
        processos = (processos_validos | processos_concluidos_meta).distinct()

        if processos.count() != len(processo_ids):
            return JsonResponse({'success': False, 'message': 'Alguns processos não são do usuário ou inválidos.'}, status=400)

        meta.meta_qtd = meta_qtd
        meta.processos.set(processos)
        meta.save()

        return JsonResponse({
            'success': True,
            'message': f'Meta atualizada com sucesso! {len(processos)} processos associados.',
        })

   

@require_http_methods(["POST"])
def excluir_meta_semanal(request):
    meta_id = request.GET.get('meta_id')
    meta = get_object_or_404(MetaSemanal, id=meta_id)
    meta.delete()
    return JsonResponse({'success': True, 'message': 'Meta excluída com sucesso!'})


@login_required
@require_http_methods(["GET"])
def ver_todos_processos_meta(request):
    """
    View para o modal "Ver Processos" que estava faltando
    """
    meta_id = request.GET.get('meta_id')
    if not meta_id:
        return JsonResponse({'error': 'meta_id é obrigatório'}, status=400)
    
    try:
        meta = get_object_or_404(MetaSemanal, id=meta_id)
        
        # Verificar permissões
        if not request.user.has_perm('app_name.view_metasemanal') and meta.usuario != request.user:
            return JsonResponse({'error': 'Sem permissão para ver esta meta'}, status=403)
        
        hoje = timezone.localdate()
        
        # Buscar processos da meta com informações detalhadas
        processos = meta.processos.select_related('especie', 'tipo').annotate(
            fase_atual=Subquery(
                ProcessoAndamento.objects.filter(
                    processo=OuterRef('pk')
                ).order_by('-dt_criacao').values('fase__fase')[:1]
            )
        )
        
        FASES_EXATAS = ["Revisão", "Revisão Des", "L. PJE", "L.PJE", "Processo Concluído"]

        # Verificar quais processos foram revisados no período da meta
        revisados_ids = set(
            ProcessoAndamento.objects.filter(
                fase__fase__in=FASES_EXATAS,
                processo__in=processos,
                dt_criacao__date__range=(meta.semana_inicio, meta.semana_fim)
            ).values_list('processo_id', flat=True)
        )
        
        # Preparar dados dos processos
        is_meta_ativa = meta.semana_fim >= hoje
        processos_data = []
        for processo in processos:
            dias_gabinete = (hoje - processo.antigo.date()).days if processo.antigo else 0
            
            is_concluido = processo.id in revisados_ids
            
            # Se a meta for da semana atual, considera a fase atual também
            if not is_concluido and is_meta_ativa:
                fase_raw = (processo.fase_atual or '')
                fase_limpa = fase_raw.strip()
                if fase_limpa in FASES_EXATAS:
                    is_concluido = True
            
            concluido = is_concluido or (processo.tipo and processo.tipo.tipo == "Monocrática" and processo.concluido)
            
            # Determinar cor do badge baseado nos dias
            if dias_gabinete <= 30:
                badge_class = "bg-success"
            elif dias_gabinete <= 40:
                badge_class = "bg-primary"
            elif dias_gabinete <= 50:
                badge_class = "bg-warning"
            else:
                badge_class = "bg-danger"
            
            tipo_nome = processo.tipo.tipo if processo.tipo else ''
            if not tipo_nome and processo.prioridade_urgente:
                tipo_nome = 'Urgentíssimo'
                
            tipo_nome_lower = tipo_nome.lower()
            if 'urgentíssimo' in tipo_nome_lower or 'urgente' in tipo_nome_lower:
                tipo_badge = 'bg-danger'
            elif 'prioridade' in tipo_nome_lower:
                tipo_badge = 'bg-warning text-dark'
            elif 'monocrática' in tipo_nome_lower:
                tipo_badge = 'bg-primary'
            elif tipo_nome:
                tipo_badge = 'bg-secondary'
            else:
                tipo_badge = ''

            processos_data.append({
                'numero_processo': processo.numero_processo,
                'especie': processo.especie.especie if processo.especie else 'N/A',
                'fase_atual': processo.fase_atual or 'Não especificado',
                'dias_gabinete': dias_gabinete,
                'badge_class': badge_class,
                'concluido': concluido,
                'status_texto': 'Concluído' if concluido else 'Pendente',
                'status_class': 'success' if concluido else 'warning',
                'tipo_nome': tipo_nome,
                'tipo_badge': tipo_badge
            })
        
        # Ordenar: concluídos primeiro, depois por dias no gabinete (decrescente)
        processos_data.sort(key=lambda x: (not x['concluido'], -x['dias_gabinete']))
        
        # Calcular estatísticas
        total_processos = len(processos_data)
        concluidos = sum(1 for p in processos_data if p['concluido'])
        pendentes = total_processos - concluidos
        progresso = round((concluidos / meta.meta_qtd * 100) if meta.meta_qtd > 0 else 0, 1)
        
        # Preparar HTML para o modal
        html_content = f"""
        <div class="row mb-3">
            <div class="col-md-3">
                <div class="text-center">
                    <h4 class="text-primary mb-1">{total_processos}</h4>
                    <small class="text-muted">Total</small>
                </div>
            </div>
            <div class="col-md-3">
                <div class="text-center">
                    <h4 class="text-success mb-1">{concluidos}</h4>
                    <small class="text-muted">Concluídos</small>
                </div>
            </div>
            <div class="col-md-3">
                <div class="text-center">
                    <h4 class="text-warning mb-1">{pendentes}</h4>
                    <small class="text-muted">Pendentes</small>
                </div>
            </div>
            <div class="col-md-3">
                <div class="text-center">
                    <h4 class="text-info mb-1">{progresso}%</h4>
                    <small class="text-muted">Progresso</small>
                </div>
            </div>
        </div>
        
        <div class="progress mb-4" style="height: 8px;">
            <div class="progress-bar bg-primary" role="progressbar" style="width: {progresso}%"></div>
        </div>
        """
        
        if processos_data:
            html_content += """
            <div class="table-responsive">
                <table class="table table-hover">
                    <thead class="table-light">
                        <tr>
                            <th>Processo</th>
                            <th>Espécie</th>
                            <th>Fase Atual</th>
                            <th>Dias no Gabinete</th>
                            <th>Tipo</th>
                            <th>Status</th>
                        </tr>
                    </thead>
                    <tbody>
            """
            
            for processo in processos_data:
                html_content += f"""
                <tr class="{'table-success' if processo['concluido'] else ''}">
                    <td class="fw-medium">{processo['numero_processo']}</td>
                    <td>{processo['especie']}</td>
                    <td>{processo['fase_atual']}</td>
                    <td>
                        <span class="badge {processo['badge_class']}">{processo['dias_gabinete']} dias</span>
                    </td>
                    <td>
                        {f'<span class="badge {processo["tipo_badge"]}">{processo["tipo_nome"]}</span>' if processo['tipo_nome'] else '-'}
                    </td>
                    <td>
                        <span class="badge bg-{processo['status_class']}">{processo['status_texto']}</span>
                    </td>
                </tr>
                """
            
            html_content += """
                    </tbody>
                </table>
            </div>
            """
        else:
            html_content += """
            <div class="text-center py-4">
                <i class="bi bi-inbox display-4 text-muted mb-3"></i>
                <h5 class="text-muted">Nenhum processo associado a esta meta</h5>
            </div>
            """
        
        return JsonResponse({
            'success': True,
            'html': html_content,
            'usuario': meta.usuario.get_full_name(),
            'periodo': f"{meta.semana_inicio.strftime('%d/%m/%Y')} a {meta.semana_fim.strftime('%d/%m/%Y')}",
            'meta_qtd': meta.meta_qtd,
            'total_processos': total_processos,
            'concluidos': concluidos,
            'pendentes': pendentes,
            'progresso': progresso
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Erro ao carregar processos: {str(e)}'}, status=500)


@login_required  
@require_http_methods(["GET"])
def api_detalhes_meta(request):
    """
    API para retornar detalhes de uma meta (para o modal de detalhes)
    """
    meta_id = request.GET.get('meta_id')
    if not meta_id:
        return JsonResponse({'error': 'meta_id é obrigatório'}, status=400)
    
    try:
        meta = get_object_or_404(MetaSemanal, id=meta_id)
        
        # Verificar permissões
        if not request.user.has_perm('app_name.view_metasemanal') and meta.usuario != request.user:
            return JsonResponse({'error': 'Sem permissão'}, status=403)
        
        hoje = timezone.localdate()
        
        # Calcular progresso
        processos_meta = Processo.objects.filter(id__in=[p.id for p in meta.processos.all()]).annotate(
            fase_atual=Subquery(
                ProcessoAndamento.objects.filter(
                    processo=OuterRef('pk')
                ).order_by('-dt_criacao').values('fase__fase')[:1]
            )
        )
        if processos_meta.exists():
            revisados_ids = set(
                ProcessoAndamento.objects.filter(
                    processo__in=processos_meta,
                    fase__fase__in=["Revisão", "Revisão Des", "Processo Concluído"],
                    dt_criacao__range=(meta.semana_inicio, meta.semana_fim)
                ).values_list('processo', flat=True)
            )
            concluidos = 0
            fases_validas = ["Revisão", "Revisão Des", "Processo Concluído"]
            for p in processos_meta:
                if (p.id in revisados_ids and p.fase_atual in fases_validas) or (p.tipo and p.tipo.tipo == "Monocrática" and p.concluido):
                    concluidos += 1
            progresso = round((concluidos / meta.meta_qtd * 100) if meta.meta_qtd > 0 else 0, 1)
        else:
            concluidos = 0
            progresso = 0
        
        # Calcular faixas de dias para o usuário
        processos_usuario = Processo.objects.filter(
            usuario=meta.usuario,
            concluido=False
        ).exclude(antigo__isnull=True)
        
        total_user = processos_usuario.count()
        if total_user > 0:
            dias = [(hoje - p.antigo.date()).days for p in processos_usuario]
            mais_30 = sum(1 for d in dias if d > 30)
            mais_40 = sum(1 for d in dias if d > 40)
            mais_50 = sum(1 for d in dias if d > 50)
            
            range_le30 = total_user - mais_30
            range_31_40 = mais_30 - mais_40
            range_41_50 = mais_40 - mais_50
            range_gt50 = mais_50
        else:
            range_le30 = range_31_40 = range_41_50 = range_gt50 = 0
        
        # Determinar status
        if meta.semana_fim < hoje:
            status = 'Passada'
            status_class = 'secondary'
        elif meta.semana_inicio > hoje:
            status = 'Futura'
            status_class = 'info'
        else:
            status = 'Atual'
            status_class = 'primary'
        
        data = {
            'success': True,
            'usuario': meta.usuario.get_full_name(),
            'periodo': f"{meta.semana_inicio.strftime('%d/%m/%Y')} a {meta.semana_fim.strftime('%d/%m/%Y')}",
            'meta_qtd': meta.meta_qtd,
            'progresso': progresso,
            'total_processos': len(processos_meta),
            'concluidos': concluidos,
            'pendentes': len(processos_meta) - concluidos,
            'range_le30': range_le30,
            'range_31_40': range_31_40,
            'range_41_50': range_41_50,
            'range_gt50': range_gt50,
            'status': status,
            'status_class': status_class
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': f'Erro ao carregar detalhes: {str(e)}'}, status=500)
    
@login_required
@require_http_methods(["GET"])
def minhas_metas(request):
    user = request.user
    agora_dt = timezone.localtime()
    hoje = agora_dt.date()
    
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    fim_semana = inicio_semana + timedelta(days=6)

    meta = MetaSemanal.objects.filter(
        usuario=user,
        semana_inicio=inicio_semana,
        semana_fim=fim_semana
    ).prefetch_related(
        Prefetch('processos', Processo.objects.select_related('especie', 'tipo'))
    ).first()

    total_meta = pendentes = concluidas = progresso = 0
    processos_com_status = []
    processos_fora_meta = []

    if meta:
        total_meta = meta.meta_qtd
        processos = list(meta.processos.all())

        concluidos_ids = set(
            ProcessoAndamento.objects.filter(
                processo__in=processos,
                fase__fase__in=["Revisão", "Revisão Des", "L. PJE", "Processo Concluído"],
                dt_criacao__date__range=(inicio_semana, fim_semana)
            ).values_list('processo', flat=True).distinct()
        )

        ultima_fase_subquery = ProcessoAndamento.objects.filter(
            processo=OuterRef('pk')
        ).order_by('-dt_criacao').values('fase__fase')[:1]
        
        processos_com_fase = Processo.objects.filter(pk__in=[p.id for p in processos]).annotate(
            fase_atual=Subquery(ultima_fase_subquery)
        )
        fases_atuais = {p.id: p.fase_atual for p in processos_com_fase}

        concluidas = 0
        fases_validas = ["Revisão", "Revisão Des", "L. PJE", "Processo Concluído"]
        for p in processos:
            fase_atual = fases_atuais.get(p.id, 'Não especificado')
            if (p.id in concluidos_ids and fase_atual in fases_validas) or (p.tipo and p.tipo.tipo == "Monocrática" and p.concluido):
                concluidas += 1

        pendentes = max(0, total_meta - concluidas)
        progresso = round((concluidas / total_meta * 100) if total_meta > 0 else 0, 1)

        for p in processos:
            data_entrada = p.antigo or p.dt_criacao
            dias_calc = 0
            if data_entrada:
                dt_ref = data_entrada.date() if hasattr(data_entrada, 'date') else data_entrada
                dias_calc = (hoje - dt_ref).days
            
            # Garante que dias_calc seja sempre inteiro
            dias_calc = int(dias_calc) if dias_calc is not None else 0

            # --- AQUI ESTAVA O ERRO: GARANTINDO QUE CONCLUIDO SEJA SEMPRE BOOL ---
            status_concluido = False
            fase_atual = fases_atuais.get(p.id, 'Não especificado')
            if p.id in concluidos_ids and fase_atual in fases_validas:
                status_concluido = True
            elif p.tipo and p.tipo.tipo == "Monocrática" and p.concluido:
                status_concluido = True
            # -------------------------------------------------------------------

            processos_com_status.append({
                'numero_processo': p.numero_processo,
                'data_entrada': data_entrada,
                'dias_no_gabinete': dias_calc,
                'fase_atual': fases_atuais.get(p.id, 'Não especificado'),
                'especie': p.especie.especie if p.especie else 'Não especificado',
                'tipo': p.tipo.tipo if p.tipo else 'Não especificado',
                'concluido': bool(status_concluido), # Força ser booleano
                'tags_materia': p.tags_materia or '',
            })

        # ORDENAÇÃO BLINDADA: Garante que os valores comparados nunca sejam None
        processos_com_status = sorted(
            processos_com_status,
            key=lambda x: (
                bool(x.get('concluido', False)), 
                -int(x.get('dias_no_gabinete', 0))
            )
        )

        # Processos fora da meta
        processos_fora_ids = list(ProcessoAndamento.objects.filter(
            processo__usuario=user,
            fase__fase="Revisão",
            dt_criacao__date__range=(inicio_semana, fim_semana)
        ).exclude(processo__in=[p.id for p in processos]).values_list('processo', flat=True).distinct())

        if processos_fora_ids:
            processos_fora_detalhes = Processo.objects.filter(
                id__in=processos_fora_ids
            ).select_related('especie', 'tipo').annotate(
                fase_atual=Subquery(ultima_fase_subquery)
            )

            for p in processos_fora_detalhes:
                data_entrada_fora = p.antigo or p.dt_criacao
                dias_fora = (hoje - data_entrada_fora.date()).days if data_entrada_fora else 0
                
                processos_fora_meta.append({
                    'numero_processo': p.numero_processo,
                    'data_entrada': data_entrada_fora,
                    'dias_no_gabinete': int(dias_fora or 0),
                    'fase_atual': p.fase_atual or 'Não especificado',
                    'especie': p.especie.especie if p.especie else 'Não especificado',
                    'tipo': p.tipo.tipo if p.tipo else 'Não especificado',
                    'concluido': True,
                    'tags_materia': p.tags_materia or '',
                })

    return render(request, 'minhas_metas.html', {
        'meta': meta,
        'total_meta': total_meta,
        'metas_pendentes': pendentes,
        'metas_concluidas': concluidas,
        'progresso': progresso,
        'processos_com_status': processos_com_status,
        'processos_fora_meta': processos_fora_meta,
    })


# ─────────────────────────────────────────────────────────────────
#  PROCESSOS EM PAUTA
# ─────────────────────────────────────────────────────────────────
from .models import ProcessoPauta
import pandas as pd
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt


@login_required
@require_POST
def importar_pauta(request):
    """
    Recebe um arquivo Excel (.xlsx), apaga toda a pauta anterior
    e importa os novos registros.
    Colunas esperadas na planilha:
      - 'Número do Processo'
      - 'Data da sessão'
      - 'Tarefa atual'  → 'Aguardando a sessão' = presencial
                          'Aguardando sessão virtual' = virtual
    """
    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        return JsonResponse({'success': False, 'message': 'Nenhum arquivo enviado.'}, status=400)
    if not arquivo.name.lower().endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'message': 'Envie um arquivo .xlsx ou .xls.'}, status=400)

    try:
        df = pd.read_excel(arquivo, engine='openpyxl')

        # Normalizar nomes das colunas (strip + title-case insensitive)
        df.columns = [str(c).strip() for c in df.columns]

        col_numero = None
        for c in df.columns:
            cl = c.lower().strip()
            if 'número' in cl or 'numero' in cl or cl == 'processo':
                col_numero = c
                break

        col_data = None
        for c in df.columns:
            cl = c.lower().strip()
            if 'data' in cl and 'sess' in cl:
                col_data = c
                break
        
        if not col_data:
            for c in df.columns:
                cl = c.lower().strip()
                if 'data' in cl and 'tarefa' not in cl and 'conclus' not in cl and 'cri' not in cl:
                    col_data = c
                    break
        
        # Aceita: "Tarefa atual", "Fase atual", "Situação", "Status", etc.
        # Exclui colunas de data (ex: "Data tarefa atual") para não confundir com a coluna de status
        col_tarefa = next((
            c for c in df.columns
            if any(kw in c.lower() for kw in ['tarefa', 'fase', 'situação', 'situacao', 'status', 'andamento'])
            and not c.lower().strip().startswith('data')
        ), None)

        if not col_numero or not col_data:
            missing = []
            if not col_numero: missing.append('Número do Processo')
            if not col_data:   missing.append('Data da sessão')
            return JsonResponse({
                'success': False,
                'message': f"Falha na leitura. Colunas achadas: {', '.join(df.columns)}. Precisa de Número e Data."
            }, status=400)

        now = timezone.localtime()

        # Apaga a pauta anterior ANTES de importar APENAS para as datas contidas na planilha.
        # Isso significa que processos de outras datas que já estavam no banco não serão perdidos.
        datas_na_planilha = set()
        
        for idx, row in df.iterrows():
            data_raw = row[col_data]
            if pd.isna(data_raw): continue
            
            if isinstance(data_raw, str):
                try:
                    d_obj = pd.to_datetime(data_raw, dayfirst=True)
                except:
                    continue
            else:
                d_obj = data_raw
            
            try:
                # O formato salvo no banco de dados contém YYYY-MM-DD
                datas_na_planilha.add(d_obj.strftime('%Y-%m-%d'))
            except:
                pass
                
        # Apaga APENAS os processos 'presencial' e 'virtual' do banco cujas datas estão na planilha.
        # Isso garante que 'terceira_camara', 'vandymara' e 'marcio_vidal' (lançados manualmente) sejam preservados.
        for dt_str in datas_na_planilha:
            ProcessoPauta.objects.filter(
                data_sessao__startswith=dt_str, 
                tipo_sessao__in=['presencial', 'virtual']
            ).delete()

        inseridos = 0
        erros = []
        cnt_virtual = 0
        cnt_presencial = 0

        for idx, row in df.iterrows():
            numero_raw = str(row[col_numero]).strip() if pd.notna(row[col_numero]) else ''
            data_raw   = row[col_data]
            tarefa_raw = str(row[col_tarefa]).strip().lower() if col_tarefa is not None and pd.notna(row[col_tarefa]) else ''

            if not numero_raw or numero_raw == 'nan':
                continue  # linha vazia

            # Tipo de sessão: se o valor contiver "virtual" → virtual, senão → presencial
            if 'virtual' in tarefa_raw:
                tipo = 'virtual'
                cnt_virtual += 1
            else:
                tipo = 'presencial'
                cnt_presencial += 1

            # Parse da data
            data_sessao = None
            if pd.notna(data_raw):
                try:
                    if isinstance(data_raw, str):
                        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                            try:
                                data_sessao = datetime.strptime(data_raw.strip(), fmt)
                                break
                            except ValueError:
                                continue
                    else:
                        data_sessao = pd.Timestamp(data_raw).to_pydatetime()
                    if data_sessao:
                        data_sessao = timezone.make_aware(data_sessao) if timezone.is_naive(data_sessao) else data_sessao
                except Exception:
                    erros.append(f"Linha {idx+2}: data inválida ({data_raw})")
                    continue

            if not data_sessao:
                erros.append(f"Linha {idx+2}: data da sessão não reconhecida.")
                continue

            # Tenta vincular ao processo no banco
            processo_db = Processo.objects.filter(numero_processo__icontains=numero_raw).first()

            ProcessoPauta.objects.create(
                numero_processo=numero_raw,
                data_sessao=data_sessao,
                tipo_sessao=tipo,
                processo_vinculado=processo_db,
            )
            inseridos += 1

        # Valores únicos na coluna tarefa (para debug)
        valores_tarefa = []
        if col_tarefa:
            valores_tarefa = list(df[col_tarefa].dropna().astype(str).str.strip().unique()[:10])

        return JsonResponse({
            'success': True,
            'message': f'{inseridos} processo(s) importado(s): {cnt_presencial} presencial(is), {cnt_virtual} virtual(is).',
            'erros': erros,
            'total': inseridos,
            'debug': {
                'col_tarefa_detectada': col_tarefa,
                'colunas_arquivo': list(df.columns),
                'valores_tarefa_amostra': valores_tarefa,
                'cnt_virtual': cnt_virtual,
                'cnt_presencial': cnt_presencial,
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao processar arquivo: {str(e)}'}, status=500)


@login_required
@require_POST
def adicionar_pauta_manual(request):
    """Adiciona um único processo à pauta manualmente."""
    import json
    if request.content_type == 'application/json':
        try:
            body = json.loads(request.body)
            numero = (body.get('numero') or body.get('numero_processo', '')).strip()
            data_str = body.get('data_sessao', '').strip()
            tipo = body.get('tipo_sessao', 'presencial').strip()
            responsavel = body.get('responsavel', '').strip()
            tema = body.get('tema', '').strip()
            especie = body.get('especie', '').strip()
            link_documento = body.get('link_documento', '').strip()
            link_memorial = body.get('link_memorial', '').strip()
        except Exception:
            return JsonResponse({'success': False, 'message': 'JSON inválido.'}, status=400)
    else:
        numero = (request.POST.get('numero') or request.POST.get('numero_processo', '')).strip()
        data_str = request.POST.get('data_sessao', '').strip()
        tipo = request.POST.get('tipo_sessao', 'presencial').strip()
        responsavel = request.POST.get('responsavel', '').strip()
        tema = request.POST.get('tema', '').strip()
        especie = request.POST.get('especie', '').strip()
        link_documento = request.POST.get('link_documento', '').strip()
        link_memorial = request.POST.get('link_memorial', '').strip()

    if not numero or not data_str:
        return JsonResponse({'success': False, 'message': 'Número do processo e data são obrigatórios.'}, status=400)

    try:
        data_sessao = datetime.strptime(data_str, '%Y-%m-%dT%H:%M')
        data_sessao = timezone.make_aware(data_sessao)
    except ValueError:
        try:
            data_sessao = datetime.strptime(data_str, '%Y-%m-%d')
            data_sessao = timezone.make_aware(data_sessao)
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Formato de data inválido.'}, status=400)

    processo_db = Processo.objects.filter(numero_processo__icontains=numero).first()

    item = ProcessoPauta.objects.create(
        numero_processo=numero,
        data_sessao=data_sessao,
        tipo_sessao=tipo if tipo in ('presencial', 'virtual', 'vandymara', 'marcio_vidal', 'terceira_camara') else 'presencial',
        processo_vinculado=processo_db,
        responsavel_manual=responsavel or None,
        tema_manual=tema or None,
        especie_manual=especie or None,
        link_documento_manual=link_documento or None,
        link_memorial=link_memorial or None,
    )
    

    return JsonResponse({
        'success': True,
        'item': {
            'id': item.id,
            'numero_processo': item.numero_processo,
            'data_sessao': item.data_sessao.strftime('%d/%m/%Y'),
            'data_sessao_hora': item.data_sessao.strftime('%H:%M'),
            'tipo_sessao': item.tipo_sessao,
            'responsavel': item.responsavel_manual or (processo_db.usuario.get_full_name() if processo_db and processo_db.usuario else 'Não atribuído'),
            'especie': item.especie_manual or (processo_db.especie.especie if processo_db and processo_db.especie else '—'),
            'tema': item.tema_manual or (processo_db.tema.nome if processo_db and processo_db.tema else '—'),
            'link_documento': item.link_documento_manual,
            'link_memorial': item.link_memorial,
            'vinculado': processo_db is not None,
        }
    })


@login_required
def pauta_json(request):
    """Retorna JSON com listas separadas de presenciais e virtuais."""
    itens = ProcessoPauta.objects.select_related(
        'processo_vinculado',
        'processo_vinculado__usuario',
        'processo_vinculado__especie',
        'processo_vinculado__tema',
    ).prefetch_related(
        'processo_vinculado__andamentos'
    ).order_by('data_sessao')

    def serializar(item):
        p = item.processo_vinculado
        link_doc = None
        if p:
            # Usa o cache do prefetch — sem query adicional por processo
            andamentos_validos = [a for a in p.andamentos.all() if a.link_doc and str(a.link_doc).strip()]
            if andamentos_validos:
                andamento_com_link = max(andamentos_validos, key=lambda x: x.dt_criacao)
                link_doc = andamento_com_link.link_doc

        responsavel_final = item.responsavel_manual or (p.usuario.get_full_name() if p and p.usuario else 'Não atribuído')
        especie_final     = item.especie_manual or (p.especie.especie if p and p.especie else '—')
        tema_final        = item.tema_manual or (p.tema.nome if p and p.tema else '—')
        link_doc_final    = item.link_documento_manual or link_doc

        return {
            'id': item.id,
            'numero_processo': item.numero_processo,
            'data_sessao': item.data_sessao.strftime('%d/%m/%Y'),
            'data_sessao_hora': item.data_sessao.strftime('%H:%M'),
            'tipo_sessao': item.tipo_sessao,
            'responsavel': responsavel_final,
            'especie': especie_final,
            'tema': tema_final,
            'vinculado': p is not None,
            'link_documento': link_doc_final,
            'link_memorial': item.link_memorial,
        }

    itens_serializados = [serializar(i) for i in itens]

    return JsonResponse({
        'itens': itens_serializados,
        'total': len(itens_serializados),
        # backward-compat para o header badge (home_metrics / desa)
        'presenciais': [x for x in itens_serializados if x['tipo_sessao'] == 'presencial'],
        'virtuais':    [x for x in itens_serializados if x['tipo_sessao'] == 'virtual'],
    })


@login_required
@require_POST
def remover_pauta_item(request, item_id):
    """Remove um único item da pauta."""
    item = get_object_or_404(ProcessoPauta, id=item_id)
    item.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def limpar_pauta(request):
    """Apaga todos os registros de ProcessoPauta."""
    total = ProcessoPauta.objects.count()
    ProcessoPauta.objects.all().delete()
    return JsonResponse({'success': True, 'message': f'{total} processo(s) removidos da pauta.'})


@login_required
@require_POST
def alterar_tipo_sessao_pauta(request, item_id):
    """Altera o tipo de sessão (presencial/virtual/vandymara/marcio_vidal) de um item da pauta."""
    import json
    item = get_object_or_404(ProcessoPauta, id=item_id)
    try:
        body = json.loads(request.body)
        novo_tipo = body.get('tipo_sessao', '')
    except Exception:
        novo_tipo = request.POST.get('tipo_sessao', '')

    tipos_validos = ('presencial', 'virtual', 'vandymara', 'marcio_vidal', 'terceira_camara')
    if novo_tipo not in tipos_validos:
        return JsonResponse({'success': False, 'message': 'Tipo inválido.'}, status=400)

    item.tipo_sessao = novo_tipo
    item.save()
    return JsonResponse({'success': True, 'tipo_sessao': item.tipo_sessao})


@login_required
@require_POST
def editar_pauta_item(request, item_id):
    """Edita os campos manuais de um item da pauta via modal de edição."""
    import json
    item = get_object_or_404(ProcessoPauta, id=item_id)
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'message': 'JSON inválido.'}, status=400)

    tipos_validos = ('presencial', 'virtual', 'vandymara', 'marcio_vidal', 'terceira_camara')
    novo_tipo = body.get('tipo_sessao', item.tipo_sessao)
    if novo_tipo not in tipos_validos:
        novo_tipo = item.tipo_sessao

    item.tipo_sessao = novo_tipo
    item.responsavel_manual = body.get('responsavel', '').strip() or None
    item.especie_manual = body.get('especie', '').strip() or None
    item.tema_manual = body.get('tema', '').strip() or None
    item.link_documento_manual = body.get('link_documento', '').strip() or None
    item.link_memorial = body.get('link_memorial', '').strip() or None
    item.save()

    return JsonResponse({'success': True})


@login_required
def avisos_lista(request):
    """Retorna o template base do Quadro de Avisos ou a lista via AJAX"""
    avisos = Aviso.objects.filter(ativo=True)
    
    # Contagem de não lidos para o badge (pode ser útil aqui também)
    avisos_nao_lidos_count = Aviso.objects.filter(ativo=True).exclude(leitores=request.user).count()
    
    context = {
        'avisos': avisos,
        'avisos_nao_lidos_count': avisos_nao_lidos_count,
        'is_chefe': request.user.profile.funcao == "Chefe de Gabinete" if hasattr(request.user, 'profile') else False
    }
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'components/avisos_lista_ajax.html', context)
        
    return render(request, 'components/avisos.html', context)


@login_required
def aviso_detalhe(request, pk):
    """Retorna o conteúdo de um aviso específico e marca como lido"""
    aviso = get_object_or_404(Aviso, pk=pk, ativo=True)
    
    # Marcar como lido
    if request.user not in aviso.leitores.all():
        aviso.leitores.add(request.user)
    
    return JsonResponse({
        'success': True,
        'titulo': aviso.titulo,
        'conteudo': aviso.conteudo,
        'autor': aviso.autor.get_full_name(),
        'data': aviso.criado_em.strftime('%d/%m/%Y %H:%M'),
        'imagem_url': aviso.imagem.url if aviso.imagem else None,
        'pdf_url': aviso.pdf.url if aviso.pdf else None,
        'fixado': aviso.fixado
    })


@login_required
@require_http_methods(["POST"])
def aviso_salvar(request):
    """Cria ou edita um aviso (Apenas Chefe de Gabinete)"""
    if not hasattr(request.user, 'profile') or request.user.profile.funcao != "Chefe de Gabinete":
        return JsonResponse({'success': False, 'message': 'Acesso negado.'}, status=403)
        
    # Suporte a multipart/form-data (POST + FILES)
    pk = request.POST.get('id')
    titulo = request.POST.get('titulo')
    conteudo = request.POST.get('conteudo')
    fixado = request.POST.get('fixado') == 'true'
    imagem = request.FILES.get('imagem')
    pdf = request.FILES.get('pdf')
    
    if pk:
        aviso = get_object_or_404(Aviso, pk=pk)
        aviso.titulo = titulo
        aviso.conteudo = conteudo
        aviso.fixado = fixado
        if imagem:
            aviso.imagem = imagem
        if pdf:
            aviso.pdf = pdf
        aviso.save()
    else:
        aviso = Aviso.objects.create(
            titulo=titulo,
            conteudo=conteudo,
            autor=request.user,
            fixado=fixado,
            imagem=imagem,
            pdf=pdf
        )
        
    return JsonResponse({'success': True, 'id': aviso.id})


@login_required
@require_http_methods(["POST"])
def aviso_deletar(request, pk):
    """Desativa um aviso (Apenas Chefe de Gabinete)"""
    if not hasattr(request.user, 'profile') or request.user.profile.funcao != "Chefe de Gabinete":
        return JsonResponse({'success': False, 'message': 'Acesso negado.'}, status=403)
        
    aviso = get_object_or_404(Aviso, pk=pk)
    aviso.ativo = False
    aviso.save()
    return JsonResponse({'success': True})

import json
from openai import OpenAI
from decouple import config

@login_required
def analisar_minuta_ia(request, processo_id):
    if request.method == 'POST':
        try:
            processo = get_object_or_404(Processo, id=processo_id)
            texto_minuta = request.POST.get('texto_minuta', '').strip()
            arquivo_recurso = request.FILES.get('arquivo_recurso')
            
            if not texto_minuta:
                return JsonResponse({'status': 'error', 'message': 'O texto da minuta não pode estar vazio.'}, status=400)
                
            if not arquivo_recurso:
                return JsonResponse({'status': 'error', 'message': 'O envio do arquivo do recurso é obrigatório para a análise da IA.'}, status=400)

            # Conectar à API da OpenAI
            api_key = config('OPENAI_API_KEY', default='')
            if not api_key or 'sk-cole' in api_key:
                return JsonResponse({'status': 'error', 'message': 'A chave OPENAI_API_KEY não foi configurada corretamente.'}, status=400)

            client = OpenAI(api_key=api_key)

            uploaded_file_id = None
            if arquivo_recurso:
                # Realiza o upload do arquivo para a OpenAI antes da chamada de IA
                response_file = client.files.create(
                    file=(arquivo_recurso.name, arquivo_recurso.read()),
                    purpose="assistants"
                )
                uploaded_file_id = response_file.id

            # Mensagem enviada ao Assistente: somente o texto da minuta.
            # O prompt/instruções vem exclusivamente do Assistente configurado na OpenAI.
            conteudo_final = f"Abaixo segue a minuta para análise:\n\n{texto_minuta}"


            # 1. Criar uma Thread
            thread = client.beta.threads.create()

            # 2. Adicionar a mensagem à Thread
            msg_params = {
                "thread_id": thread.id,
                "role": "user",
                "content": conteudo_final
            }
            if uploaded_file_id:
                msg_params["attachments"] = [
                    {
                        "file_id": uploaded_file_id,
                        "tools": [{"type": "file_search"}]
                    }
                ]
            
            client.beta.threads.messages.create(**msg_params)

            # 3. Executar o Assistente
            try:
                import time
                run = client.beta.threads.runs.create(
                    thread_id=thread.id,
                    assistant_id="asst_YgwStqtIr9VrXrXeVl6QiWFc"
                )

                # 4. Aguardar a conclusão (Polling)
                while run.status in ['queued', 'in_progress', 'cancelling']:
                    time.sleep(1.5)
                    run = client.beta.threads.runs.retrieve(
                        thread_id=thread.id,
                        run_id=run.id
                    )

                if run.status == 'completed':
                    # 5. Resgatar a resposta final
                    messages = client.beta.threads.messages.list(
                        thread_id=thread.id
                    )
                    resposta_ia = messages.data[0].content[0].text.value
                else:
                    raise Exception(f"A execução do assistente falhou ou não foi finalizada corretamente. Status: {run.status}")
            finally:
                # Exclui o arquivo do Storage da OpenAI para economizar espaço
                if uploaded_file_id:
                    try:
                        client.files.delete(uploaded_file_id)
                    except Exception:
                        pass

            texto_formatado = f"🤖 **Análise Prévia da IA:**\n\n{resposta_ia}"
            ComentarioProcesso.objects.create(
                processo=processo,
                usuario=request.user,
                texto=texto_formatado
            )
            
            return JsonResponse({'status': 'success', 'message': 'Análise concluída e adicionada aos comentários.'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erro na IA: {str(e)}'}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método inválido'}, status=400)


# ══════════════════════════════════════════════════════════════════
#  MÓDULO: GESTÃO DE FÉRIAS E PLANTÕES
# ══════════════════════════════════════════════════════════════════

from .models import Ferias, Plantao, NotificacaoInterna
import json
from django.core.exceptions import ValidationError as DjangoValidationError


def _is_chefe_gabinete(user):
    """Verifica se o usuário tem perfil de Chefe de Gabinete."""
    try:
        return user.profile.funcao == 'Chefe de Gabinete'
    except Exception:
        return user.is_staff


# ── View principal ────────────────────────────────────────────────

@login_required
def gestao_ferias_plantoes(request):
    """Renderiza a página do Gráfico de Gantt de Férias e Plantões."""
    usuarios = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    is_chefe = _is_chefe_gabinete(request.user)
    return render(request, 'processos/gestao_ferias_plantoes.html', {
        'usuarios': usuarios,
        'is_chefe': is_chefe,
    })


# ── APIs JSON para o Gantt ────────────────────────────────────────

@login_required
def ferias_json(request):
    """Retorna férias em formato JSON para vis-timeline."""
    hoje = timezone.localdate()
    
    # Atualiza automaticamente para 'em_andamento' caso a data tenha chegado
    Ferias.objects.filter(
        status='aprovado',
        data_inicio__lte=hoje,
        data_fim__gte=hoje
    ).update(status='em_andamento')

    ferias_qs = Ferias.objects.all().select_related('usuario')

    COR_STATUS = {
        'pendente':    {'background': '#f59e0b', 'border': '#d97706'},
        'aprovado':    {'background': '#10b981', 'border': '#059669'},
        'em_andamento':{'background': '#3b82f6', 'border': '#2563eb'},
        'cancelado':   {'background': '#9ca3af', 'border': '#6b7280'},
    }

    items = []
    groups = {}
    for f in ferias_qs:
        uid = f.usuario_id
        if uid not in groups:
            groups[uid] = {
                'id': f'u{uid}',
                'content': f.usuario.get_full_name() or f.usuario.username,
            }
        cor = COR_STATUS.get(f.status, COR_STATUS['pendente'])
        data_label = f'{f.data_inicio.strftime("%d/%m/%Y")} – {f.data_fim.strftime("%d/%m/%Y")}'
        items.append({
            'id': f'ferias-{f.pk}',
            'group': f'u{uid}',
            'content': f'<span title="{f.get_status_display()}">🏖 {data_label}</span>',
            'start': f.data_inicio.isoformat(),
            'end': f.data_fim.isoformat(),
            'style': f'background-color:{cor["background"]};border-color:{cor["border"]};color:#fff;border-radius:4px;font-size:0.8rem;padding:2px 6px;',
            'type': 'range',
            'pk': f.pk,
            'status': f.status,
            'status_label': f.get_status_display(),
            'observacoes': f.observacoes,
            'periodo_aquisicao': f.periodo_aquisicao,
            'usuario_id': uid,
            'usuario_nome': f.usuario.get_full_name(),
            'data_label': data_label,
        })

    return JsonResponse({'groups': list(groups.values()), 'items': items})


@login_required
def plantoes_json(request):
    """Retorna plantões em formato JSON para vis-timeline."""
    hoje = timezone.localdate()
    
    # Atualiza automaticamente para 'em_andamento' caso a data tenha chegado
    Plantao.objects.filter(
        status='confirmado',
        data_inicio__lte=hoje,
        data_fim__gte=hoje
    ).update(status='em_andamento')

    plantoes_qs = Plantao.objects.all().select_related('usuario')

    COR_STATUS = {
        'pendente':    {'background': '#f59e0b', 'border': '#d97706'},
        'confirmado':  {'background': '#10b981', 'border': '#059669'},
        'em_andamento':{'background': '#3b82f6', 'border': '#2563eb'},
        'cancelado':   {'background': '#9ca3af', 'border': '#6b7280'},
    }

    items = []
    groups = {}
    meu_id = request.user.pk

    for p in plantoes_qs:
        uid = p.usuario_id
        if uid not in groups:
            groups[uid] = {
                'id': f'u{uid}',
                'content': p.usuario.get_full_name() or p.usuario.username,
            }
        cor = COR_STATUS.get(p.status, COR_STATUS['pendente'])
        pode_confirmar = (uid == meu_id and p.status == 'pendente')
        data_label = f'{p.data_inicio.strftime("%d/%m/%Y")} – {p.data_fim.strftime("%d/%m/%Y")}'
        items.append({
            'id': f'plantao-{p.pk}',
            'group': f'u{uid}',
            'content': f'<span title="{p.get_status_display()}">⚖️ {data_label}</span>',
            'start': p.data_inicio.isoformat(),
            'end': p.data_fim.isoformat(),
            'style': f'background-color:{cor["background"]};border-color:{cor["border"]};color:#fff;border-radius:4px;font-size:0.8rem;padding:2px 6px;',
            'type': 'range',
            'pk': p.pk,
            'status': p.status,
            'status_label': p.get_status_display(),
            'observacoes': p.observacoes,
            'usuario_id': uid,
            'usuario_nome': p.usuario.get_full_name(),
            'data_label': data_label,
            'pode_confirmar': pode_confirmar,
        })

    return JsonResponse({'groups': list(groups.values()), 'items': items})


# ── CRUD Férias ───────────────────────────────────────────────────

@login_required
@require_http_methods(["POST"])
def ferias_criar(request):
    if not _is_chefe_gabinete(request.user):
        return JsonResponse({'success': False, 'message': 'Acesso restrito ao Chefe de Gabinete.'}, status=403)
    try:
        data = json.loads(request.body)
        usuario = get_object_or_404(User, pk=data['usuario_id'])
        
        confirmado = data.get('confirmado', False)
        if not confirmado:
            qs_global = Ferias.objects.filter(
                status__in=['pendente', 'aprovado', 'em_andamento'],
                data_inicio__lte=data['data_fim'],
                data_fim__gte=data['data_inicio'],
            ).exclude(usuario=usuario)
            if qs_global.exists():
                conflito = qs_global.select_related('usuario').first()
                msg = (f"Atenção: {conflito.usuario.get_full_name()} já está com férias agendadas de "
                       f"{conflito.data_inicio.strftime('%d/%m/%Y')} a "
                       f"{conflito.data_fim.strftime('%d/%m/%Y')}. Deseja prosseguir com o agendamento mesmo assim?")
                return JsonResponse({'success': False, 'conflito': True, 'message': msg})

        ferias = Ferias(
            usuario=usuario,
            data_inicio=data['data_inicio'],
            data_fim=data['data_fim'],
            status=data.get('status', 'pendente'),
            observacoes=data.get('observacoes', ''),
            periodo_aquisicao=data.get('periodo_aquisicao', ''),
            criado_por=request.user,
        )
        ferias.full_clean()
        ferias.save(skip_validation=True)
        return JsonResponse({'success': True, 'message': 'Férias registradas com sucesso!', 'pk': ferias.pk})
    except DjangoValidationError as e:
        erros = e.message_dict if hasattr(e, 'message_dict') else {'erro': e.messages}
        return JsonResponse({'success': False, 'message': erros}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ferias_editar(request, pk):
    if not _is_chefe_gabinete(request.user):
        return JsonResponse({'success': False, 'message': 'Acesso restrito ao Chefe de Gabinete.'}, status=403)
    ferias = get_object_or_404(Ferias, pk=pk)
    try:
        data = json.loads(request.body)
        
        confirmado = data.get('confirmado', False)
        if not confirmado:
            data_inicio = data.get('data_inicio', ferias.data_inicio)
            data_fim = data.get('data_fim', ferias.data_fim)
            qs_global = Ferias.objects.filter(
                status__in=['pendente', 'aprovado', 'em_andamento'],
                data_inicio__lte=data_fim,
                data_fim__gte=data_inicio,
            ).exclude(usuario=ferias.usuario).exclude(pk=pk)
            if qs_global.exists():
                conflito = qs_global.select_related('usuario').first()
                msg = (f"Atenção: {conflito.usuario.get_full_name()} já está com férias agendadas de "
                       f"{conflito.data_inicio.strftime('%d/%m/%Y')} a "
                       f"{conflito.data_fim.strftime('%d/%m/%Y')}. Deseja prosseguir com a edição mesmo assim?")
                return JsonResponse({'success': False, 'conflito': True, 'message': msg})

        ferias.data_inicio = data.get('data_inicio', ferias.data_inicio)
        ferias.data_fim = data.get('data_fim', ferias.data_fim)
        ferias.status = data.get('status', ferias.status)
        ferias.observacoes = data.get('observacoes', ferias.observacoes)
        ferias.periodo_aquisicao = data.get('periodo_aquisicao', ferias.periodo_aquisicao)
        ferias.full_clean()
        ferias.save(skip_validation=True)
        return JsonResponse({'success': True, 'message': 'Férias atualizadas com sucesso!'})
    except DjangoValidationError as e:
        erros = e.message_dict if hasattr(e, 'message_dict') else {'erro': e.messages}
        return JsonResponse({'success': False, 'message': erros}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ferias_deletar(request, pk):
    if not _is_chefe_gabinete(request.user):
        return JsonResponse({'success': False, 'message': 'Acesso restrito ao Chefe de Gabinete.'}, status=403)
    ferias = get_object_or_404(Ferias, pk=pk)
    ferias.delete()
    return JsonResponse({'success': True, 'message': 'Férias excluídas.'})


# ── CRUD Plantão ──────────────────────────────────────────────────

@login_required
@require_http_methods(["POST"])
def plantao_criar(request):
    if not _is_chefe_gabinete(request.user):
        return JsonResponse({'success': False, 'message': 'Acesso restrito ao Chefe de Gabinete.'}, status=403)
    try:
        data = json.loads(request.body)
        usuario = get_object_or_404(User, pk=data['usuario_id'])
        plantao = Plantao(
            usuario=usuario,
            data_inicio=data['data_inicio'],
            data_fim=data['data_fim'],
            status=data.get('status', 'pendente'),
            observacoes=data.get('observacoes', ''),
            criado_por=request.user,
        )
        plantao.full_clean()
        plantao.save(skip_validation=True)
        return JsonResponse({'success': True, 'message': 'Plantão escalado com sucesso!', 'pk': plantao.pk})
    except DjangoValidationError as e:
        erros = e.message_dict if hasattr(e, 'message_dict') else {'erro': e.messages}
        return JsonResponse({'success': False, 'message': erros}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def plantao_editar(request, pk):
    if not _is_chefe_gabinete(request.user):
        return JsonResponse({'success': False, 'message': 'Acesso restrito ao Chefe de Gabinete.'}, status=403)
    plantao = get_object_or_404(Plantao, pk=pk)
    try:
        data = json.loads(request.body)
        plantao.data_inicio = data.get('data_inicio', plantao.data_inicio)
        plantao.data_fim = data.get('data_fim', plantao.data_fim)
        plantao.status = data.get('status', plantao.status)
        plantao.observacoes = data.get('observacoes', plantao.observacoes)
        plantao.full_clean()
        plantao.save(skip_validation=True)
        return JsonResponse({'success': True, 'message': 'Plantão atualizado com sucesso!'})
    except DjangoValidationError as e:
        erros = e.message_dict if hasattr(e, 'message_dict') else {'erro': e.messages}
        return JsonResponse({'success': False, 'message': erros}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def plantao_deletar(request, pk):
    if not _is_chefe_gabinete(request.user):
        return JsonResponse({'success': False, 'message': 'Acesso restrito ao Chefe de Gabinete.'}, status=403)
    plantao = get_object_or_404(Plantao, pk=pk)
    plantao.delete()
    return JsonResponse({'success': True, 'message': 'Plantão excluído.'})


# ── Confirmação de Ciência (assessor) ────────────────────────────

@login_required
@require_http_methods(["POST"])
def plantao_confirmar_ciencia(request, pk):
    """O próprio assessor confirma ciência do plantão. Status: pendente → confirmado."""
    plantao = get_object_or_404(Plantao, pk=pk)

    if plantao.usuario != request.user:
        return JsonResponse(
            {'success': False, 'message': 'Você só pode confirmar ciência dos seus próprios plantões.'},
            status=403
        )
    if plantao.status != 'pendente':
        return JsonResponse(
            {'success': False, 'message': f'Este plantão já está com status "{plantao.get_status_display()}".'},
            status=400
        )

    Plantao.objects.filter(pk=pk).update(status='confirmado')
    NotificacaoInterna.objects.filter(
        destinatario=request.user, tipo='plantao', lida=False, link='/gestao/ferias-plantoes/',
    ).update(lida=True)
    return JsonResponse({'success': True, 'message': 'Ciência confirmada! Status atualizado para Confirmado.'})


@login_required
@require_http_methods(["POST"])
def ferias_confirmar_ciencia(request, pk):
    """O próprio assessor confirma ciência das suas férias. Status: pendente → aprovado."""
    ferias = get_object_or_404(Ferias, pk=pk)

    if ferias.usuario != request.user:
        return JsonResponse(
            {'success': False, 'message': 'Você só pode confirmar ciência das suas próprias férias.'},
            status=403
        )
    if ferias.status not in ('pendente', 'em_andamento'):
        return JsonResponse(
            {'success': False, 'message': f'Estas férias já estão com status "{ferias.get_status_display()}".'},
            status=400
        )

    Ferias.objects.filter(pk=pk).update(status='aprovado')
    NotificacaoInterna.objects.filter(
        destinatario=request.user, tipo='ferias', lida=False, link='/gestao/ferias-plantoes/',
    ).update(lida=True)
    return JsonResponse({'success': True, 'message': 'Ciência das férias confirmada! Status: Aprovado.'})


# ── Notificações ──────────────────────────────────────────────────

@login_required
def notificacoes_json(request):
    """Retorna notificações não lidas do usuário atual (para badge/polling)."""
    notifs = NotificacaoInterna.objects.filter(
        destinatario=request.user, lida=False
    ).values('id', 'tipo', 'titulo', 'mensagem', 'link', 'criado_em')

    resultado = []
    for n in notifs:
        n['criado_em'] = n['criado_em'].strftime('%d/%m/%Y %H:%M')
        resultado.append(n)

    return JsonResponse({'count': len(resultado), 'notificacoes': resultado})


@login_required
@require_http_methods(["POST"])
def notificacao_marcar_lida(request, pk):
    """Marca uma notificação como lida."""
    notif = get_object_or_404(NotificacaoInterna, pk=pk, destinatario=request.user)
    notif.lida = True
    notif.save(update_fields=['lida'])
    return JsonResponse({'success': True})
