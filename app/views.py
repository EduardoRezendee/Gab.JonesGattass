
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from django.core.cache import cache
from django.db.models import Count, Q
from django.http import JsonResponse
from django.db.models.functions import Left
# MODELS
from processos.models import Processo, TarefaDoDia, ComentarioProcesso, ProcessoAndamento, Tema, Tipo, Aviso
from accounts.models import UserProfile
from django.utils import timezone
from django.http import JsonResponse
from django.db.models import Count
from datetime import datetime, timedelta, timezone as dt_timezone
from processos.models import ProcessoAndamento, User
from django.db.models import F, ExpressionWrapper, DateField, DurationField, IntegerField, CharField, Value, Max
from django.db.models.functions import Cast
from collections import defaultdict
from processos.models import ProcessoAndamento


def home3(request):
    return render(request, 'home3.html')
# MÉTRICAS E GAMIFICAÇÃO
from .metrics import (
    get_process_metrics,
    get_process_gamification_metrics,
    get_top_users_by_xp,
    get_pending_and_concluded_by_assessor,
    get_daily_entries_and_exits_by_assessor,
    get_user_daily_productivity,
    get_user_weekly_productivity,
    get_user_meta_semanal_metrics
)

@login_required(login_url='login')
def home(request):
    user = request.user
    
    atrasados_por_assessor_detalhado: list[dict] = []
    total_atrasados: int = 0

    # Verificar roles do usuário (cache simples)
    cache_key_roles = f'user_roles_{user.id}'
    roles = cache.get(cache_key_roles)
    if not roles:
        roles = {
            'is_revisor': UserProfile.objects.filter(user=user, funcao="revisor(a)").exists(),
            'is_desembargador': UserProfile.objects.filter(user=user, funcao="Desembargador").exists(),
            'is_chefe': UserProfile.objects.filter(user=user, funcao="Chefe de Gabinete").exists(),
            'is_assessor': UserProfile.objects.filter(user=user, funcao="Assessor(a)").exists(),
        }
        cache.set(cache_key_roles, roles, timeout=3600)

    is_revisor = roles['is_revisor']
    is_desembargador = roles['is_desembargador']
    is_chefe = roles['is_chefe']
    is_assessor = roles['is_assessor']

    # Inicializar variáveis
    andamento_metrics = []
    numero_de_processos_em_revisao = 0
    numero_de_processos_em_revisao_des = 0
    active_tab = None
    fases = None
    hoje = timezone.now()
    processos_nao_concluidos = Processo.objects.none()
    processos_detalhados = []  # Inicializar aqui para evitar UnboundLocalError
    processos_antigos_detalhados = []
    processos_liminares_detalhados = []
    processos_concluidos_detalhados = []
    processos_revisao_des_detalhados = []
    tarefas_detalhadas = []
    tarefas_ids = []
    numero_de_processos_em_revisao_des = 0
    fase_filtro = None
    numero_processo = request.GET.get('numero_processo', '').strip()


    # Métricas diárias
    revisoes_hoje = ProcessoAndamento.objects.filter(
        fase__fase__in=["Revisão", "Revisão Des"],
        dt_criacao__date=hoje.date()
    ).values('processo').distinct().count()

    concluidos_revisao_hoje = ProcessoAndamento.objects.filter(
        fase__fase__in=["Concluído", "Revisão Des"],
        dt_conclusao__date=hoje.date(),
        status__status="Concluído"
    ).values('processo').distinct().count()

    # Métricas gerais (com cache)
    cache_key_metrics = 'general_metrics'
    general_metrics = cache.get(cache_key_metrics)
    if not general_metrics:
        total_pendentes = Processo.objects.filter(concluido=False).count()
        processos_por_fase = (
            ProcessoAndamento.objects.filter(
                processo__concluido=False,
                status__status__in=["Não iniciado", "Em andamento"]
            )
            .values('fase__fase')
            .annotate(quantidade=Count('processo', distinct=True))
            .order_by('fase__fase')
        )
        general_metrics = {
            'total_pendentes': total_pendentes,
            'processos_por_fase': processos_por_fase
        }
        cache.set(cache_key_metrics, general_metrics, timeout=3600)

    total_pendentes = general_metrics['total_pendentes']
    processos_por_fase = general_metrics['processos_por_fase']

    # VISÃO DO REVISOR(A)
    if is_revisor:
        processos_em_revisao = Processo.objects.filter(
            andamentos__fase__fase="Revisão",
            andamentos__usuario=user,
            concluido=False
        ).distinct().select_related('especie', 'usuario').prefetch_related('andamentos', 'andamentos__fase', 'andamentos__status', 'andamentos__usuario')
        if numero_processo:
            processos_em_revisao = processos_em_revisao.filter(numero_processo__icontains=numero_processo)
        
        comentarios_dict = {
            p.pk: list(ComentarioProcesso.objects.filter(processo=p).select_related('usuario'))
            for p in processos_em_revisao
        }
        for processo in processos_em_revisao:
            total_comentarios = processo.comentarios.count()
            ultimo_andamento = processo.andamentos.filter(
                fase__fase="Revisão",
                usuario=user,
                status__status__in=["Não iniciado", "Em andamento"]
            ).order_by('-dt_criacao').first()
            if ultimo_andamento:
                especie_nome = processo.especie.especie if processo.especie else "Sem espécie"
                comentarios = comentarios_dict.get(processo.pk, [])
                andamento_metrics.append({
                    'pk': ultimo_andamento.pk,
                    'processo_pk': processo.pk,
                    'numero_processo': processo.numero_processo,
                    'data_dist': processo.data_dist,
                    'dias_no_gabinete': processo.dias_no_gabinete() or 0,
                    'fase': ultimo_andamento.fase.fase,
                    'descricao': ultimo_andamento.andamento,
                    'status': ultimo_andamento.status.status,
                    'data_inicio': ultimo_andamento.dt_inicio,
                    'data_conclusao': ultimo_andamento.dt_conclusao,
                    'tipo': processo.tipo.tipo if processo.tipo else "Sem tipo",
                    'link_doc': ultimo_andamento.link_doc,
                    'usuario_processo': processo.usuario.get_full_name() if processo.usuario else "Não atribuído",
                    'especie': especie_nome,
                    'sigla_especie': processo.especie.sigla if processo.especie else "",
                    'data_envio_revisao': ultimo_andamento.dt_criacao,
                    'tema': processo.tema.nome if processo.tema else None,
                    'comentarios': [{'texto': c.texto, 'data_criacao': c.data_criacao, 'usuario': c.usuario.get_full_name()} for c in comentarios],
                    'tem_comentario': total_comentarios > 0, # <-- ADICIONE ESTA LINHA
                    'total_comentarios': total_comentarios,   # <-- ADICIONE ESTA LINHA

                })
        andamento_metrics.sort(
                    key=lambda p: (0 if p.get('tipo') == "2ª Correção" else 1,
                        0 if p.get('tipo') == "Redistribuído" else 2,
                        0 if p.get('especie') == "Voto Vista" else 3,
                        0 if p.get('especie') == "EDCIV" else 4,
                        0 if p.get('especie') == "Revisitado" else 5,
                        0 if p.get('especie') == "Voto de Divergência" else 6,
                        0 if p.get('tipo') == "Urgentíssimo" else 7,
                        0 if p.get('tipo') == "Prioridade" else 8,
                        0 if p.get('tipo') == "Monocrática" else 9,                      
                        -(p.get('dias_no_gabinete') or 0))
)


 # --- VISÃO DA DESEMBARGADORA ---

    elif is_desembargador:
        fase_filtro = request.GET.get('fase', 'Revisão Des')
        if fase_filtro not in ['Revisão Des', 'Devolvido', 'Revisão']:
            fase_filtro = 'Revisão Des'

        # Pegamos somente processos com último andamento na fase selecionada.
        # Se for 'Devolvido' ou 'Revisão', mostramos visão GERAL (de todos os assessores).
        # Se for 'Revisão Des', mostramos apenas o que está com o Desembargador logado.
        query_revisao = Q(concluido=False, andamentos__fase__fase=fase_filtro)
        if fase_filtro not in ['Devolvido', 'Revisão']:
            query_revisao &= Q(andamentos__usuario=user)

        processos_em_revisao_des = (
            Processo.objects
            .filter(query_revisao)
            .annotate(
                max_dt_criacao=Max('andamentos__dt_criacao')
            )
            .filter(
                andamentos__dt_criacao=F('max_dt_criacao'),
                andamentos__status__status__in=["Não iniciado", "Em andamento"]
            )
            .distinct()
            .select_related('especie', 'usuario', 'tipo')
            .prefetch_related('andamentos', 'andamentos__fase', 'andamentos__status', 'andamentos__usuario')
        )

        if numero_processo:
            processos_em_revisao_des = processos_em_revisao_des.filter(numero_processo__icontains=numero_processo)

        numero_de_processos_em_revisao_des = processos_em_revisao_des.count()

        comentarios_dict = {
            p.pk: list(ComentarioProcesso.objects.filter(processo=p).select_related('usuario'))
            for p in processos_em_revisao_des
        }

        andamento_metrics = []
        for processo in processos_em_revisao_des:
            # Último andamento em Revisão Des desta desembargadora (em aberto)
            total_comentarios = processo.comentarios.count()
            # último andamento na fase filtrada (em aberto)
            filt_andamento = Q(fase__fase=fase_filtro, status__status__in=["Não iniciado", "Em andamento"])
            if fase_filtro not in ['Devolvido', 'Revisão']:
                filt_andamento &= Q(usuario=user)

            ultimo_andamento = processo.andamentos.filter(filt_andamento).order_by('-dt_criacao').first()

            if ultimo_andamento:
                especie_nome = processo.especie.especie if processo.especie else "Sem espécie"
                tipo_nome = processo.tipo.tipo if processo.tipo else "Sem tipo"
                comentarios = comentarios_dict.get(processo.pk, [])
                revisoes_des_count = processo.andamentos.filter(fase__fase="Revisão Des").count()

                andamento_metrics.append({
                    'pk': ultimo_andamento.pk,
                    'processo_pk': processo.pk,
                    'numero_processo': processo.numero_processo,
                    'data_dist': processo.data_dist,
                    'dias_no_gabinete': processo.dias_no_gabinete() or 0,
                    'fase': ultimo_andamento.fase.fase,
                    'descricao': ultimo_andamento.andamento,
                    'status': ultimo_andamento.status.status,
                    'data_inicio': ultimo_andamento.dt_inicio,
                    'data_conclusao': ultimo_andamento.dt_conclusao,
                    'link_doc': ultimo_andamento.link_doc,
                    'usuario_processo': processo.usuario.get_full_name() if processo.usuario else "Não atribuído",
                    'especie': especie_nome,
                    'sigla_especie': processo.especie.sigla if processo.especie else "",
                    'tipo': tipo_nome,
                    'comentarios': [
                        {'texto': c.texto, 'data_criacao': c.data_criacao, 'usuario': c.usuario.get_full_name()}
                        for c in comentarios
                    ],
                    'revisoes_des': revisoes_des_count,
                    'data_envio_revisao_des': ultimo_andamento.dt_criacao,
                    'tema': processo.tema.nome if processo.tema else None,
                    'camara': processo.camara.camara if getattr(processo, 'camara', None) else "Sem câmara",
                    'tem_comentario': total_comentarios > 0, # <-- ADICIONE ESTA LINHA
                    'total_comentarios': total_comentarios,   # <-- ADICIONE ESTA LINHA
                })

        # Ordenação: Plantão primeiro, depois Liminar, depois mais tempo no gabinete
        andamento_metrics.sort(
            key=lambda p: (0 if p.get('tipo') == "2ª Correção" else 1,
                        0 if p.get('tipo') == "Redistribuído" else 2,
                        0 if p.get('especie') == "Voto Vista" else 3,
                        0 if p.get('especie') == "EDCIV" else 4,
                        0 if p.get('especie') == "Revisitado" else 5,
                        0 if p.get('especie') == "Voto de Divergência" else 6,
                        0 if p.get('tipo') == "Urgentíssimo" else 7,
                        0 if p.get('tipo') == "Prioridade" else 8,
                        0 if p.get('tipo') == "Monocrática" else 9,                      
                        -(p.get('dias_no_gabinete') or 0))
        )

        # Lista de mais antigos
        processos_mais_antigos = (
            Processo.objects.filter(concluido=False, antigo__isnull=False)
            .select_related('especie', 'usuario', 'tipo')
            .order_by('antigo')[:10]
        )
        processos_antigos_detalhados = []
        for processo in processos_mais_antigos:
            ultimo_andamento = processo.andamentos.order_by('-dt_criacao').first()
            dias_no_gabinete = (hoje - processo.antigo).days if processo.antigo else 0
            tipo_nome = processo.tipo.tipo if processo.tipo else "Sem tipo"
            processos_antigos_detalhados.append({
                'numero_processo': processo.numero_processo,
                'especie': processo.especie.especie if processo.especie else "Sem espécie",
                'data_entrada_gabinete': processo.antigo,
                'dias_no_gabinete': dias_no_gabinete,
                'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem fase",
                'usuario': processo.usuario.get_full_name() if processo.usuario else "Não atribuído",
                'tipo': tipo_nome,
            })
        processos_antigos_detalhados.sort(
            key=lambda p: (0 if p.get('tipo') == "2ª Correção" else 1,
                        0 if p.get('tipo') == "Redistribuído" else 2,
                        0 if p.get('especie') == "Voto Vista" else 3,
                        0 if p.get('especie') == "EDCIV" else 4,
                        0 if p.get('especie') == "Revisitado" else 5,
                        0 if p.get('especie') == "Voto de Divergência" else 6,
                        0 if p.get('tipo') == "Urgentíssimo" else 7,
                        0 if p.get('tipo') == "Prioridade" else 8,
                        0 if p.get('tipo') == "Monocrática" else 9,                      
                        -(p.get('dias_no_gabinete') or 0))
        )

        # Lista de liminares
        processos_liminares = (
            Processo.objects.filter(concluido=False, especie__especie="Liminar")
            .select_related('especie', 'usuario', 'tipo')
            .order_by('antigo')
        )
        processos_liminares_detalhados = []
        for processo in processos_liminares:
            ultimo_andamento = processo.andamentos.order_by('-dt_criacao').first()
            dias_no_gabinete = (hoje - processo.data_dist).days if processo.data_dist else 0
            tipo_nome = processo.tipo.tipo if processo.tipo else "Sem tipo"
            processos_liminares_detalhados.append({
                'numero_processo': processo.numero_processo,
                'data_entrada_gabinete': processo.data_dist,
                'dias_no_gabinete': dias_no_gabinete,
                'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem fase",
                'usuario': processo.usuario.get_full_name() if processo.usuario else "Não atribuído",
                'tipo': tipo_nome,
            })
        processos_liminares_detalhados.sort(
            key=lambda p: (0 if p['tipo'] == "Prioridade" else 1, -(p['dias_no_gabinete'] or 0))
        )

    # VISÃO DO CHEFE DE GABINETE
    elif is_chefe:
        andamento_metrics = []
        processos_mais_antigos = Processo.objects.filter(
            concluido=False,
            antigo__isnull=False
        ).select_related('especie', 'usuario').order_by('antigo')[:10]
        processos_antigos_detalhados = []
        for processo in processos_mais_antigos:
            ultimo_andamento = processo.andamentos.order_by('-dt_criacao').first()
            dias_no_gabinete = (hoje - processo.antigo).days if processo.antigo else 0
            processos_antigos_detalhados.append({
                'numero_processo': processo.numero_processo,
                'especie': processo.especie.especie if processo.especie else "Sem espécie",
                'data_entrada_gabinete': processo.antigo,
                'dias_no_gabinete': dias_no_gabinete,
                'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem fase",
                'usuario': processo.usuario.get_full_name() if processo.usuario else "Não atribuído",
            })
        processos_liminares = Processo.objects.filter(
            concluido=False,
            especie__especie="Liminar"
        ).select_related('especie', 'usuario').order_by('antigo')
        processos_liminares_detalhados = []
        for processo in processos_liminares:
            ultimo_andamento = processo.andamentos.order_by('-dt_criacao').first()
            dias_no_gabinete = (hoje - processo.data_dist).days if processo.data_dist else 0
            processos_liminares_detalhados.append({
                'numero_processo': processo.numero_processo,
                'data_entrada_gabinete': processo.data_dist,
                'dias_no_gabinete': dias_no_gabinete,
                'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem fase",
                'usuario': processo.usuario.get_full_name() if processo.usuario else "Não atribuído",
            })
             # Processos com +30 dias por assessor
        processos_mais_30 = (
            Processo.objects
            .filter(concluido=False, antigo__isnull=False, usuario__isnull=False)
            .annotate(
                antigo_date=Cast('antigo', output_field=DateField()),
                dias_no_gabinete=ExpressionWrapper(
                    timezone.now().date() - F('antigo_date'),
                    output_field=DurationField()
                )
            )
            .filter(dias_no_gabinete__gt=timedelta(days=30))
            .select_related('usuario', 'especie', 'tipo')
            .prefetch_related('andamentos', 'andamentos__fase')
        )

        # Agrupar por assessor com contagem
        atrasados_por_assessor = defaultdict(list)
        for p in processos_mais_30:
            nome_assessor = p.usuario.get_full_name()
            ultimo_andamento = p.andamentos.order_by('-dt_criacao').first()
            atrasados_por_assessor[nome_assessor].append({
                'numero_processo': p.numero_processo,
                'dias_no_gabinete': p.dias_no_gabinete.days,
                'especie': p.especie.sigla if p.especie else '—',
                'tipo': p.tipo.tipo if p.tipo else 'Sem tipo',
                'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else 'Sem fase',
                'is_liminar': p.tipo.tipo == "Liminar" if p.tipo else False
            })

        # Criar lista estruturada para o template
        atrasados_por_assessor_detalhado = [
            {
                'assessor': nome,
                'quantidade': len(processos),
                'processos': sorted(
                    processos,
                    key=lambda x: (
                        0 if x['is_liminar'] else 1,  # Priorizar liminares
                        -x['dias_no_gabinete']  # Maior tempo no gabinete
                    )
                )
            }
            for nome, processos in atrasados_por_assessor.items()
        ]
        
        # Ordenar assessores por quantidade de processos atrasados (descendente) ou nome
        atrasados_por_assessor_detalhado.sort(key=lambda x: (-x['quantidade'], x['assessor']))

        # Quantitativo total
        total_atrasados = processos_mais_30.count()

    # VISÃO DO ASSESSOR/USUÁRIO COMUM
    else:
        despacho = request.GET.get('despacho', '').strip()
        tipo = request.GET.get('tipo', '').strip()
        especie = request.GET.get('especie', '').strip()  # Novo parâmetro para espécie

        processos_nao_concluidos = Processo.objects.filter(
            usuario=user,
            concluido=False
        ).select_related('especie', 'usuario').prefetch_related('andamentos', 'andamentos__fase', 'andamentos__status')

        # Aplicar filtros
        if numero_processo:
            numero_processo = numero_processo.strip()[:10]
            processos_nao_concluidos = processos_nao_concluidos.annotate(
                numero_processo_truncado=Left('numero_processo', 10)
            ).filter(
                numero_processo_truncado__icontains=numero_processo
            )
        if despacho:
            if despacho.lower() == 'sim':
                processos_nao_concluidos = processos_nao_concluidos.filter(despacho=True)
            elif despacho.lower() == 'nao':
                processos_nao_concluidos = processos_nao_concluidos.filter(despacho=False)
        if tipo:
            processos_nao_concluidos = processos_nao_concluidos.filter(tipo__tipo__icontains=tipo)
        if especie:  # Novo filtro por espécie
            processos_nao_concluidos = processos_nao_concluidos.filter(especie__sigla__iexact=especie)

        if not is_revisor and not is_desembargador and not is_chefe:
            comentarios_dict = {
                p.pk: list(ComentarioProcesso.objects.filter(processo=p).select_related('usuario'))
                for p in processos_nao_concluidos
            }
            for processo in processos_nao_concluidos:
                total_comentarios = processo.comentarios.count()
                ultimo_andamento = processo.andamentos.order_by('-dt_criacao').first()
                if ultimo_andamento and processo.pk:
                    processos_detalhados.append({
                        'pk': processo.pk,
                        'andamento_pk': ultimo_andamento.pk,
                        'andamento_link_doc': ultimo_andamento.link_doc if ultimo_andamento else None,
                        'numero_processo': processo.numero_processo,
                        'especie': processo.especie.especie if processo.especie else "Sem espécie",
                        'sigla': processo.especie.sigla if processo.especie else "Sem sigla",
                        'tema': processo.tema.nome if processo.tema else None,
                        'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem fase",
                        'status_atual': ultimo_andamento.status.status if ultimo_andamento and ultimo_andamento.status else "Sem status",
                        'data_prazo': processo.dt_prazo,
                        'data_dist': processo.data_dist,
                        'despacho': processo.despacho,
                        'prioridade_urgente': processo.prioridade_urgente,
                        'tipo': processo.tipo.tipo if processo.tipo else "Sem tipo",
                        'dias_no_gabinete': processo.dias_no_gabinete() or 0,
                        'usuario_processo': processo.usuario.get_full_name() if processo.usuario else "Não atribuído",
                        'comentarios': [
                            {'texto': c.texto, 'data_criacao': c.data_criacao, 'usuario': c.usuario.get_full_name()}
                            for c in comentarios_dict.get(processo.pk, [])],
                        'tem_comentario': total_comentarios > 0, # <-- ADICIONE ESTA LINHA
                        'total_comentarios': total_comentarios,   # <-- ADICIONE ESTA LINHA
                        
                    })
            processos_detalhados.sort(
                key=lambda p: (0 if p.get('tipo') == "2ª Correção" else 1,
                            0 if p.get('tipo') == "Redistribuído" else 2,
                            0 if p.get('especie') == "Voto Vista" else 3,
                            0 if p.get('especie') == "EDCIV" else 4,
                            0 if p.get('especie') == "Revisitado" else 5,
                            0 if p.get('especie') == "Voto de Divergência" else 6,
                            0 if p.get('tipo') == "Urgentíssimo" else 7,
                            0 if p.get('tipo') == "Prioridade" else 8,
                            0 if p.get('tipo') == "Monocrática" else 9,                      
                            -(p.get('dias_no_gabinete') or 0))
        )

            fixed_phase_order = ['Elaboração', 'Revisão', 'Correção', 'Revisão Des', 'Devolvido', 'L. PJE']
            phase_dict = {}
            for processo in processos_detalhados:
                phase_dict.setdefault(processo['fase_atual'], []).append(processo)

            fases = [(phase, phase_dict.get(phase, [])) for phase in fixed_phase_order if phase in phase_dict]
            fase_param = request.GET.get('fase', None)
            if fase_param and fase_param in fixed_phase_order and fase_param in phase_dict:
                active_tab = fase_param
            elif fases:
                active_tab = fases[0][0]

        # Nova Lista: Processos Concluídos do Usuário Logado (apenas hoje)
        processos_concluidos = Processo.objects.filter(
            usuario=user,
            concluido=True,
            dt_conclusao__date=hoje
        ).select_related('especie', 'usuario').prefetch_related('andamentos', 'andamentos__fase', 'andamentos__status')
        for processo in processos_concluidos:
            ultimo_andamento = processo.andamentos.order_by('-dt_criacao').first()
            if ultimo_andamento and processo.pk:
                processos_concluidos_detalhados.append({
                    'pk': processo.pk,
                    'andamento_pk': ultimo_andamento.pk,
                    'numero_processo': processo.numero_processo,
                    'especie': processo.especie.especie if processo.especie else "Sem espécie",
                    'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem fase",
                    'data_dist': processo.data_dist,
                    'dt_conclusao': processo.dt_conclusao,
                    'dias_no_gabinete': processo.dias_no_gabinete() or 0,
                })
        processos_concluidos_detalhados.sort(
            key=lambda p: p['dt_conclusao'] if p['dt_conclusao'] is not None else timezone.make_aware(timezone.datetime.min),
            reverse=True
        )

        # Nova Lista: Processos Enviados para Revisão Desa pelo Usuário Logado (apenas hoje)
        processos_revisao_des = Processo.objects.filter(
            usuario=user,
            andamentos__fase__fase="Revisão Des",
            andamentos__dt_criacao__date=hoje
        ).distinct().select_related('especie', 'usuario').prefetch_related('andamentos', 'andamentos__fase', 'andamentos__status')
        for processo in processos_revisao_des:
            ultimo_andamento = processo.andamentos.order_by('-dt_criacao').first()
            andamento_revisao_des = processo.andamentos.filter(
                fase__fase="Revisão Des",
                dt_criacao__date=hoje
            ).order_by('-dt_criacao').first()
            if ultimo_andamento and andamento_revisao_des and processo.pk:
                processos_revisao_des_detalhados.append({
                    'pk': processo.pk,
                    'andamento_pk': ultimo_andamento.pk,
                    'numero_processo': processo.numero_processo,
                    'especie': processo.especie.especie if processo.especie else "Sem espécie",
                    'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem fase",
                    'data_dist': processo.data_dist,
                    'data_envio_revisao_des': andamento_revisao_des.dt_criacao,
                    'dias_no_gabinete': processo.dias_no_gabinete() or 0,
                })
        processos_revisao_des_detalhados.sort(
            key=lambda p: p['data_envio_revisao_des'] if p['data_envio_revisao_des'] is not None else timezone.make_aware(timezone.datetime.min),
            reverse=True
        )

        # Adicionar espécies disponíveis para o formulário de filtros
        especies = Processo.objects.filter(
            usuario=user,
            concluido=False
        ).values('especie__sigla').distinct()
    # Tarefas do Dia (comum a todos os papéis)
    tarefas_do_dia = TarefaDoDia.objects.filter(usuario=user).select_related('processo', 'processo__especie', 'processo__tema').prefetch_related('processo__andamentos', 'processo__andamentos__fase')
    for tarefa in tarefas_do_dia:
        ultimo_andamento = tarefa.processo.andamentos.order_by('-dt_criacao').first() if tarefa.processo else None
        tarefa_dict = {
            'id': tarefa.id,
            'processo': {
                'id': tarefa.processo.id if tarefa.processo else None,
                'numero_processo': tarefa.processo.numero_processo if tarefa.processo else "Sem número",
                'especie': tarefa.processo.especie.especie if tarefa.processo and tarefa

.processo.especie else "Sem espécie",
                'fase_atual': ultimo_andamento.fase.fase if ultimo_andamento and ultimo_andamento.fase else "Sem fase",
                'data_dist': tarefa.processo.data_dist if tarefa.processo else None,
                'tema': tarefa.processo.tema.nome if tarefa.processo and tarefa.processo.tema else "Sem tema",
                'dias_no_gabinete': tarefa.processo.dias_no_gabinete() if tarefa.processo else 0,
                'dt_prazo': tarefa.processo.dt_prazo if tarefa.processo else None,
                'andamento_pk': ultimo_andamento.pk if ultimo_andamento else None,
                'andamento_link_doc': ultimo_andamento.link_doc if ultimo_andamento else None,
                'andamento': ultimo_andamento,
                'comentarios': ComentarioProcesso.objects.filter(processo=tarefa.processo).select_related('usuario') if tarefa.processo else []
            }
        }
        tarefas_detalhadas.append(tarefa_dict)
    tarefas_ids = [tarefa['processo']['id'] for tarefa in tarefas_detalhadas if tarefa['processo']['id']]
    
    # Processos de todas as espécies com mais de 30 dias
    processos_queryset = (
        Processo.objects
        .filter(usuario=user, concluido=False, antigo__isnull=False)
        .annotate(
            antigo_date = Cast('antigo', output_field=DateField()),
            dias_no_gabinete = ExpressionWrapper(
                timezone.now().date() - F('antigo_date'), output_field=DurationField()
            )
        )
        .filter(dias_no_gabinete__gt=timedelta(days=30))
        .select_related('especie')
    )

    processos_atrasados = [
        {
            "numero_processo": p.numero_processo,
            "dias_no_gabinete": p.dias_no_gabinete.days,
            "especie": p.especie.sigla,
        }
        for p in processos_queryset
    ]

    # Métricas e Gamificação (comum a todos os papéis)
    process_metrics = get_process_metrics(user)
    if not is_revisor and not is_desembargador and not is_chefe:
        process_metrics['detalhes_processos'] = processos_detalhados
    process_gamification = get_process_gamification_metrics(user)
    meta_semanal_progresso = get_user_meta_semanal_metrics(user)
    top_users = get_top_users_by_xp()
    photo_url = user.profile.photo.url if hasattr(user, 'profile') and user.profile.photo else '/static/default-profile.png'

    # Contexto comum a todos os papéis
    context = {
        'user': user,
        'is_revisor': is_revisor,
        'is_desembargador': is_desembargador,
        'is_chefe': is_chefe,
        'is_assessor': is_assessor,
        'andamento_metrics': andamento_metrics,
        'process_metrics': process_metrics,
        'process_gamification': process_gamification,
        'meta_semanal_progresso': meta_semanal_progresso,
        'top_users': top_users,
        'photo_url': photo_url,
        'tarefas_ids': tarefas_ids,
        'tarefas_do_dia': tarefas_detalhadas,
        'fases': fases,
        'active_tab': active_tab,
        'temas': Tema.objects.all(),
        'today': hoje,
        'tipos': Tipo.objects.all(),
        'revisoes_hoje': revisoes_hoje,
        'concluidos_revisao_hoje': concluidos_revisao_hoje,
        'total_pendentes': total_pendentes,
        'processos_por_fase': processos_por_fase,
        'processos_antigos_detalhados': processos_antigos_detalhados,
        'processos_liminares_detalhados': processos_liminares_detalhados,
        'processos_concluidos_detalhados': processos_concluidos_detalhados,
        'processos_revisao_des_detalhados': processos_revisao_des_detalhados,
        'numero_de_processos_em_revisao_des': numero_de_processos_em_revisao_des,
        "processos_atrasados": processos_atrasados,
        'atrasados_por_assessor': atrasados_por_assessor_detalhado,
        'total_atrasados': total_atrasados,
        'fase_ativa_desa': fase_filtro,
        'avisos_nao_lidos_count': Aviso.objects.filter(ativo=True).exclude(leitores=user).count(),
    }

    if not is_revisor and not is_desembargador and not is_chefe:
        context['show_productivity_charts'] = False

    return render(request, 'home.html', context)

# Endpoints para Chart.js
def get_pending_concluded_data(request):
    data = get_pending_and_concluded_by_assessor()
    return JsonResponse(data)

def get_entries_exits_data(request):
    data = get_daily_entries_and_exits_by_assessor()
    return JsonResponse(data)

def get_revisoes_hoje_data(request):
    # Obter a data atual
    hoje = timezone.now()
    
    # Definir o intervalo de datas: início e fim do dia atual
    data_inicio = timezone.make_aware(timezone.datetime(hoje.year, hoje.month, hoje.day, 0, 0, 0))
    data_fim = timezone.make_aware(timezone.datetime(hoje.year, hoje.month, hoje.day, 23, 59, 59))

    # Contar processos enviados para "Revisão Desa" hoje, agrupados por assessor responsável pelo processo
    enviados_por_assessor = ProcessoAndamento.objects.filter(
        fase__fase="Revisão Des",
        dt_criacao__range=(data_inicio, data_fim)
    ).values('processo__usuario__first_name', 'processo__usuario__last_name').annotate(
        total=Count('processo', distinct=True)
    ).order_by('processo__usuario__first_name', 'processo__usuario__last_name')

    # Contar o total de processos enviados para "Revisão Desa" hoje
    total_enviados = ProcessoAndamento.objects.filter(
        fase__fase="Revisão Des",
        dt_criacao__range=(data_inicio, data_fim)
    ).values('processo').distinct().count()

    # Preparar os dados para o gráfico
    labels = [f"{assessor['processo__usuario__first_name']} {assessor['processo__usuario__last_name']}" for assessor in enviados_por_assessor]
    labels.append("Total Enviados")  # Adicionar o total como último rótulo

    data = [assessor['total'] for assessor in enviados_por_assessor]
    data.append(total_enviados)  # Adicionar o total como último valor

    # Gerar cores dinamicamente para cada assessor (e o total)
    cores = ['#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6', '#EC4899', '#6B7280']  # Lista de cores
    background_colors = []
    for i in range(len(labels)):
        background_colors.append(cores[i % len(cores)])  # Ciclar pelas cores se houver mais assessores

    # Atualizar o rótulo para refletir apenas o dia atual
    response_data = {
        'labels': labels,
        'datasets': [{
            'label': f'Enviados para Revisão Des ({hoje.strftime("%d/%m/%Y")})',
            'data': data,
            'backgroundColor': background_colors
        }]
    }
    return JsonResponse(response_data)

def get_fases_data(request):
    total_pendentes = Processo.objects.filter(concluido=False).count()
    processos_por_fase = (
        ProcessoAndamento.objects.filter(
            processo__concluido=False,
            status__status__in=["Não iniciado", "Em andamento"]
        )
        .values('fase__fase')
        .annotate(quantidade=Count('processo', distinct=True))
        .order_by('fase__fase')
    )

    fases_nomes = [f['fase__fase'] for f in processos_por_fase]
    fases_quantidades = [f['quantidade'] for f in processos_por_fase]

    data = {
        'labels': fases_nomes + ['Total Pendentes'],
        'datasets': [
            {
                'label': 'Por Fase',
                'data': fases_quantidades + [total_pendentes],
                'backgroundColor': ['#3B82F6'] * len(fases_nomes) + ['#EF4444']
            }
        ]
    }
    return JsonResponse(data)

def get_es_assessor_hoje_data(request):
    hoje = timezone.now().date()
    entradas_hoje = (
        Processo.objects.filter(data_dist__date=hoje, usuario__isnull=False)
        .values('usuario__first_name', 'usuario__last_name')
        .annotate(quantidade=Count('id'))
    )
    saidas_hoje = (
        Processo.objects.filter(dt_conclusao__date=hoje, concluido=True, usuario__isnull=False)
        .values('usuario__first_name', 'usuario__last_name')
        .annotate(quantidade=Count('id'))
    )

    # Criar uma lista única de nomes de assessores a partir de entradas e saídas
    assessores_nomes = list(set(
        [f"{e['usuario__first_name']} {e['usuario__last_name']}".strip() for e in entradas_hoje] +
        [f"{s['usuario__first_name']} {s['usuario__last_name']}".strip() for s in saidas_hoje]
    ))

    entradas_dict = {f"{e['usuario__first_name']} {e['usuario__last_name']}".strip(): e['quantidade'] for e in entradas_hoje}
    saidas_dict = {f"{s['usuario__first_name']} {s['usuario__last_name']}".strip(): s['quantidade'] for s in saidas_hoje}
    entradas_vals = [entradas_dict.get(nome, 0) for nome in assessores_nomes]
    saidas_vals = [saidas_dict.get(nome, 0) for nome in assessores_nomes]

    data = {
        'labels': assessores_nomes,
        'datasets': [
            {
                'label': 'Entradas',
                'data': entradas_vals,
                'backgroundColor': '#3B82F6'
            },
            {
                'label': 'Saídas',
                'data': saidas_vals,
                'backgroundColor': '#F59E0B'
            }
        ]
    }
    return JsonResponse(data)

def get_especies_data(request):
    cache_key_especies = 'processos_por_especie'
    processos_por_especie = cache.get(cache_key_especies)
    if not processos_por_especie:
        processos_por_especie = (
            Processo.objects.filter(concluido=False)
            .values('especie__especie')
            .annotate(quantidade=Count('id'))
            .order_by('especie__especie')
        )
        cache.set(cache_key_especies, processos_por_especie, timeout=3600)

    especies_nomes = [e['especie__especie'] if e['especie__especie'] else "Sem Espécie" for e in processos_por_especie]
    especies_quantidades = [e['quantidade'] for e in processos_por_especie]

    data = {
        'labels': especies_nomes,
        'datasets': [{
            'label': 'Processos por Espécie',
            'data': especies_quantidades,
            'backgroundColor': '#10B981'
        }]
    }
    return JsonResponse(data)

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import os

@login_required
def change_profile_photo(request):
    if request.method == 'POST':
        if 'photo' in request.FILES:
            user_profile, created = UserProfile.objects.get_or_create(user=request.user)
            # Deletar a foto antiga, se existir
            if user_profile.photo and os.path.isfile(user_profile.photo.path):
                os.remove(user_profile.photo.path)
            # Salvar a nova foto
            user_profile.photo = request.FILES['photo']
            user_profile.save()
            messages.success(request, 'Foto de perfil atualizada com sucesso!')
            return redirect('home')
        else:
            messages.error(request, 'Por favor, selecione uma imagem.')
    return render(request, 'change_profile_photo.html', {'user_profile': request.user.profile})


def get_user_weekly_productivity_data(request):
    if request.user.is_authenticated:
        data = get_user_weekly_productivity(request.user)
        return JsonResponse(data)
    return JsonResponse({'error': 'Usuário não autenticado'}, status=403)

def get_user_daily_productivity_data(request):
    if request.user.is_authenticated:
        data = get_user_daily_productivity(request.user)
        return JsonResponse(data)
    return JsonResponse({'error': 'Usuário não autenticado'}, status=403)

def get_revisoes_hoje_data(request):
    hoje = timezone.localtime(timezone.now())
    data_str = request.GET.get('data', hoje.strftime('%Y-%m-%d'))

    try:
        data = datetime.strptime(data_str, '%Y-%m-%d')
    except ValueError:
        data = hoje

    local_tz = timezone.get_current_timezone()
    local_data_inicio = timezone.make_aware(datetime(data.year, data.month, data.day, 0, 0, 0), local_tz)
    local_data_fim = timezone.make_aware(datetime(data.year, data.month, data.day, 23, 59, 59), local_tz)

    data_inicio = local_data_inicio.astimezone(dt_timezone.utc)  # Usar dt_timezone.utc
    data_fim = local_data_fim.astimezone(dt_timezone.utc)  # Usar dt_timezone.utc

    enviados_por_assessor = ProcessoAndamento.objects.filter(
        fase__fase="Revisão Des",
        dt_criacao__range=(data_inicio, data_fim),
        processo__usuario__isnull=False
    ).values(
        'processo__usuario__id',
        'processo__usuario__first_name',
        'processo__usuario__last_name'
    ).annotate(
        enviados_des=Count('processo', distinct=True)
    ).order_by('processo__usuario__first_name', 'processo__usuario__last_name')

    meta = 5
    labels, data_enviados, photo_urls = [], [], []

    for assessor in enviados_por_assessor:
        user_id = assessor['processo__usuario__id']
        first_name = assessor['processo__usuario__first_name'] or 'Sem Nome'
        last_name = assessor['processo__usuario__last_name'] or ''
        enviados = assessor['enviados_des']

        labels.append(f"{first_name} {last_name}")
        data_enviados.append(enviados)

        try:
            user = User.objects.get(id=user_id)
            photo_url = (
                user.profile.photo.url if hasattr(user, 'profile') and user.profile.photo
                else f"https://via.placeholder.com/60/083464/FFFFFF?text={first_name[0].upper()}"
            )
        except User.DoesNotExist:
            photo_url = f"https://via.placeholder.com/60/083464/FFFFFF?text={first_name[0].upper()}"

        photo_urls.append(photo_url)

    data_meta = [meta] * len(labels)

    response_data = {
        'labels': labels,
        'datasets': [
            {'label': f'Enviados para Revisão Des ({data.strftime("%d/%m/%Y")})', 'data': data_enviados, 'photo_urls': photo_urls},
            {'label': 'Meta', 'data': data_meta}
        ]
    }
    return JsonResponse(response_data)


from django.utils import timezone
from django.http import JsonResponse
from django.db.models import Count
from datetime import datetime, timedelta, timezone as dt_timezone 
from processos.models import ProcessoAndamento, User

def get_revisoes_semana_data(request):
    hoje = timezone.localtime(timezone.now())
    inicio_semana = hoje - timedelta(days=hoje.weekday())  # segunda-feira
    fim_semana = inicio_semana + timedelta(days=6)  # domingo

    local_tz = timezone.get_current_timezone()
    local_data_inicio = timezone.make_aware(datetime(inicio_semana.year, inicio_semana.month, inicio_semana.day, 0, 0, 0), local_tz)
    local_data_fim = timezone.make_aware(datetime(fim_semana.year, fim_semana.month, fim_semana.day, 23, 59, 59), local_tz)

    data_inicio = local_data_inicio.astimezone(dt_timezone.utc)  # Usar dt_timezone.utc
    data_fim = local_data_fim.astimezone(dt_timezone.utc)  # Usar dt_timezone.utc

    enviados_por_assessor = ProcessoAndamento.objects.filter(
        fase__fase="Revisão Des",
        dt_criacao__range=(data_inicio, data_fim),
        processo__usuario__isnull=False  # Evita processos sem usuário
    ).values(
        'processo__usuario__id',
        'processo__usuario__first_name',
        'processo__usuario__last_name'
    ).annotate(
        enviados_des=Count('processo', distinct=True)
    ).order_by('processo__usuario__first_name')

    labels, data_enviados, photo_urls, medias_diarias = [], [], [], []

    for item in enviados_por_assessor:
        user_id = item['processo__usuario__id']
        first_name = item['processo__usuario__first_name'] or 'Sem Nome'
        last_name = item['processo__usuario__last_name'] or ''
        enviados = item['enviados_des']

        media_dia = enviados / 4  # média considerando 5 dias úteis (segunda a sexta)
        media_dia = round(media_dia, 2)  # arredonda para 2 casas decimais

        try:
            user = User.objects.get(id=user_id)
            photo_url = (
                user.profile.photo.url if hasattr(user, 'profile') and user.profile.photo
                else '/static/default-profile.png'
            )
        except User.DoesNotExist:
            photo_url = '/static/default-profile.png'

        labels.append(f"{first_name} {last_name}")
        data_enviados.append(enviados)
        photo_urls.append(photo_url)
        medias_diarias.append(media_dia)

    response_data = {
        'labels': labels,
        'datasets': [
            {'label': 'Enviados para Revisão na Semana', 'data': data_enviados, 'photo_urls': photo_urls},
            {'label': 'Média por Dia', 'data': medias_diarias}
        ]
    }
    return JsonResponse(response_data)

import json
import os
from openai import OpenAI
from django.conf import settings
from decouple import config
from django.views.decorators.csrf import csrf_exempt

@login_required
def chat_ia_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            question = data.get('question', '')
            
            if not question:
                return JsonResponse({'response': 'Por favor, faça uma pergunta válida.'})

            # Conectar à API da OpenAI
            api_key = config('OPENAI_API_KEY', default='')
            if not api_key or 'sk-cole' in api_key:
                return JsonResponse({'response': 'A chave OPENAI_API_KEY não foi configurada corretamente. Verifique o seu .env.'})
                
            assistant_id = config('OPENAI_ASSISTANT_ID', default='')
            if not assistant_id:
                return JsonResponse({'response': 'O código do Assistente (OPENAI_ASSISTANT_ID) não foi adicionado ao arquivo .env ainda.'})
                
            client = OpenAI(api_key=api_key)
            
            # 1. Cria uma "Sala de Conversa" no lado da OpenAI e manda a Pergunta do Assessor
            thread = client.beta.threads.create()
            client.beta.threads.messages.create(
                thread_id=thread.id,
                role="user",
                content=question
            )
            
            # 2. Acorda o "Assistente de Gabinete" e manda ele consultar os 5 mil PDFs e responder
            run = client.beta.threads.runs.create_and_poll(
                thread_id=thread.id,
                assistant_id=assistant_id
            )
            
            # 3. Pega a resposta final
            if run.status == 'completed':
                messages = client.beta.threads.messages.list(
                    thread_id=thread.id
                )
                resposta_ia = messages.data[0].content[0].text.value
                
                # Opcional: A OpenAI costuma adicionar referências feias [^1^] ou 【4:0†source】. Limpa via Regex.
                import re
                resposta_limpa = re.sub(r'【.*?】|\[\^.*?\^\]', '', resposta_ia)
                
                return JsonResponse({'response': resposta_limpa})
            else:
                detalhe_erro = run.last_error.message if hasattr(run, 'last_error') and run.last_error else run.status
                return JsonResponse({'response': f'A OpenAI bloqueou a resposta. Motivo: {detalhe_erro}'})
            
        except Exception as e:
            return JsonResponse({'response': f'Erro no processamento da IA: {str(e)}'})
    
    return JsonResponse({'error': 'Apenas requisições POST são permitidas'}, status=400)


# ─── AGENDA DO DESEMBARGADOR ──────────────────────────────────────────────────

from processos.models import Compromisso
import json

def _is_agenda_authorized(request):
    """Verifica se o usuário é Desembargador ou Chefe de Gabinete."""
    user = request.user
    return UserProfile.objects.filter(
        user=user,
        funcao__in=["Desembargador", "Chefe de Gabinete"]
    ).exists()


@login_required(login_url='login')
def agenda_eventos_json(request):
    """Retorna todos os compromissos como JSON, opcionalmente filtrado por mês/ano."""
    if not _is_agenda_authorized(request):
        return JsonResponse({'error': 'Acesso negado'}, status=403)

    mes = request.GET.get('mes')
    ano = request.GET.get('ano')

    qs = Compromisso.objects.all()
    if mes and ano:
        try:
            qs = qs.filter(data__month=int(mes), data__year=int(ano))
        except ValueError:
            pass

    eventos = []
    for c in qs:
        eventos.append({
            'id': c.pk,
            'titulo': c.titulo,
            'tipo': c.tipo,
            'data': c.data.strftime('%Y-%m-%d'),
            'hora_inicio': c.hora_inicio.strftime('%H:%M') if c.hora_inicio else '',
            'hora_fim': c.hora_fim.strftime('%H:%M') if c.hora_fim else '',
            'local': c.local,
            'descricao': c.descricao,
            'cor': c.cor,
            'presencial': c.presencial,
            'numero_processo': c.numero_processo,
            'cancelado': c.cancelado,
            'criado_por': c.criado_por.get_full_name() if c.criado_por else '',
        })
    return JsonResponse({'eventos': eventos})


@login_required(login_url='login')
def agenda_criar(request):
    """Cria um novo compromisso."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not _is_agenda_authorized(request):
        return JsonResponse({'error': 'Acesso negado'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    try:
        c = Compromisso.objects.create(
            titulo=data.get('titulo', '').strip(),
            tipo=data.get('tipo', 'geral'),
            data=data['data'],
            hora_inicio=data['hora_inicio'],
            hora_fim=data.get('hora_fim') or None,
            local=data.get('local', ''),
            descricao=data.get('descricao', ''),
            cor=data.get('cor', '#083464'),
            presencial=data.get('presencial', True),
            numero_processo=data.get('numero_processo', ''),
            criado_por=request.user,
        )
        return JsonResponse({'ok': True, 'id': c.pk})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required(login_url='login')
def agenda_editar(request, pk):
    """Edita um compromisso existente."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not _is_agenda_authorized(request):
        return JsonResponse({'error': 'Acesso negado'}, status=403)

    try:
        c = Compromisso.objects.get(pk=pk)
    except Compromisso.DoesNotExist:
        return JsonResponse({'error': 'Compromisso não encontrado'}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    c.titulo = data.get('titulo', c.titulo).strip()
    c.tipo = data.get('tipo', c.tipo)
    c.data = data.get('data', c.data)
    c.hora_inicio = data.get('hora_inicio', c.hora_inicio)
    c.hora_fim = data.get('hora_fim') or None
    c.local = data.get('local', c.local)
    c.descricao = data.get('descricao', c.descricao)
    c.cor = data.get('cor', c.cor)
    c.presencial = data.get('presencial', c.presencial)
    c.numero_processo = data.get('numero_processo', c.numero_processo)
    c.save()
    return JsonResponse({'ok': True})


@login_required(login_url='login')
def agenda_excluir(request, pk):
    """Exclui um compromisso."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not _is_agenda_authorized(request):
        return JsonResponse({'error': 'Acesso negado'}, status=403)

    try:
        Compromisso.objects.get(pk=pk).delete()
        return JsonResponse({'ok': True})
    except Compromisso.DoesNotExist:
        return JsonResponse({'error': 'Compromisso não encontrado'}, status=404)


@login_required(login_url='login')
def agenda_importar_bookings(request):
    """
    Importa atendimentos a partir de um arquivo TSV/CSV/XLSX exportado do Microsoft Bookings.
    Colunas esperadas (case-insensitive, flexível):
      - Data/Date, Hora/Start time, Nome/Customer name, Presencial (Sim/Não), Número do processo/Process
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    if not _is_agenda_authorized(request):
        return JsonResponse({'error': 'Acesso negado'}, status=403)

    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        return JsonResponse({'error': 'Nenhum arquivo enviado'}, status=400)

    nome = arquivo.name.lower()
    importados = 0
    erros = []

    try:
        import io
        if nome.endswith('.tsv'):
            import csv
            content = arquivo.read().decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(content), delimiter='\t')
            rows = list(reader)
        elif nome.endswith('.csv'):
            import csv
            content = arquivo.read().decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif nome.endswith('.xlsx') or nome.endswith('.xls'):
            import openpyxl

            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            return JsonResponse({'error': 'Formato não suportado. Use .csv ou .xlsx'}, status=400)

        def _find_col(row_dict, candidates):
            """Encontra chave no dicionário de forma flexível (case-insensitive, strip)."""
            for key in row_dict:
                if key is None:
                    continue
                key_lower = str(key).lower().strip()
                for c in candidates:
                    if c in key_lower:
                        return row_dict[key]
            return None

        from datetime import date as date_type, time as time_type, datetime as datetime_type
        import json as _json

        for i, row in enumerate(rows, start=2):
            try:
                # ── Date Time (coluna combinada do Bookings: "04/03/2026 15:00") ──
                raw_dt = _find_col(row, ['date time', 'data e hora', 'datetime'])
                if raw_dt is None:
                    # Fallback: tenta colunas separadas
                    raw_dt = _find_col(row, ['data', 'date', 'start date'])
                if raw_dt is None:
                    continue

                hora_from_dt = None

                if isinstance(raw_dt, datetime_type):
                    evento_data = raw_dt.date()
                    hora_from_dt = raw_dt.time()
                elif isinstance(raw_dt, date_type):
                    evento_data = raw_dt
                else:
                    raw_dt = str(raw_dt).strip()
                    # Tenta data+hora (formato do Bookings)
                    parsed_dt = None
                    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S',
                                '%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S',
                                '%m/%d/%Y %H:%M', '%m/%d/%Y %H:%M:%S'):
                        try:
                            parsed_dt = datetime_type.strptime(raw_dt, fmt)
                            break
                        except ValueError:
                            continue
                    if parsed_dt:
                        evento_data = parsed_dt.date()
                        hora_from_dt = parsed_dt.time()
                    else:
                        # Converte data pura
                        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y'):
                            try:
                                evento_data = datetime_type.strptime(raw_dt, fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            erros.append(f"Linha {i}: data inválida '{raw_dt}'")
                            continue

                # ── Hora início ──
                hora_inicio = hora_from_dt
                if hora_inicio is None:
                    raw_hora = _find_col(row, ['hora', 'start time', 'horário', 'time'])
                    if isinstance(raw_hora, time_type):
                        hora_inicio = raw_hora
                    elif isinstance(raw_hora, datetime_type):
                        hora_inicio = raw_hora.time()
                    else:
                        try:
                            hora_inicio = datetime_type.strptime(str(raw_hora or '08:00').strip()[:5], '%H:%M').time()
                        except ValueError:
                            hora_inicio = datetime_type.strptime('08:00', '%H:%M').time()

                # ── Hora fim ──
                raw_hora_fim = _find_col(row, ['duration', 'duração', 'end time', 'término', 'end'])
                hora_fim = None
                if raw_hora_fim:
                    if isinstance(raw_hora_fim, time_type):
                        hora_fim = raw_hora_fim
                    elif isinstance(raw_hora_fim, datetime_type):
                        hora_fim = raw_hora_fim.time()
                    else:
                        try:
                            hora_fim = datetime_type.strptime(str(raw_hora_fim).strip()[:5], '%H:%M').time()
                        except ValueError:
                            hora_fim = None

                # ── Nome do cliente (Customer Name) ──
                titulo = ''
                raw_nome = _find_col(row, ['customer name', 'nome', 'client name', 'name'])
                if raw_nome:
                    titulo = str(raw_nome).strip()
                if not titulo:
                    titulo = 'Atendimento'

                # ── Custom Fields (JSON do Bookings) ──
                # O Bookings exporta um campo "Custom Fields" com JSON:
                # {"Reunião Presencial?": "Sim", "Número do Processo:": "XXXX"}
                presencial = True
                numero_processo = ''

                raw_cf = _find_col(row, ['custom fields', 'custom field', 'campos personalizados'])
                if raw_cf:
                    try:
                        cf = _json.loads(str(raw_cf))
                        # Presencial
                        pres_val = cf.get('Reunião Presencial?', cf.get('presencial', cf.get('Presencial', '')))
                        if str(pres_val).strip().lower() in ('não', 'nao', 'no', 'false', '0', 'virtual'):
                            presencial = False
                        # Número do processo
                        proc_val = cf.get('Número do Processo:', cf.get('Número do Processo', cf.get('numero_processo', '')))
                        numero_processo = str(proc_val).strip() if proc_val else ''
                        if numero_processo.lower() in ('none', 'nan', 'n/i', ''):
                            numero_processo = ''
                    except Exception:
                        pass  # campo não é JSON válido — usa defaults

                # Fallback presencial se não veio em Custom Fields
                if not raw_cf:
                    raw_pres = _find_col(row, ['presencial', 'location type', 'modalidade'])
                    if raw_pres is not None:
                        if str(raw_pres).strip().lower() in ('não', 'nao', 'no', 'virtual', 'online', 'false', '0'):
                            presencial = False

                # Fallback nº processo se não veio em Custom Fields
                if not numero_processo:
                    raw_proc = _find_col(row, ['número do processo', 'processo', 'process number', 'proc'])
                    if raw_proc:
                        numero_processo = str(raw_proc).strip()
                        if numero_processo.lower() in ('none', 'nan', 'n/i', ''):
                            numero_processo = ''

                # ── Local ──
                local_raw = _find_col(row, ['sala', 'room', 'local', 'address', 'endereço'])
                local = str(local_raw).strip() if local_raw else ''
                if local.lower() in ('none', 'nan'):
                    local = ''

                existente = Compromisso.objects.filter(
                    titulo__iexact=titulo[:200],
                    data=evento_data,
                    hora_inicio=hora_inicio
                ).first()

                if existente:
                    # Atualiza os dados do existente (atualiza com os novos)
                    existente.hora_fim = hora_fim
                    existente.local = local[:200]
                    existente.presencial = presencial
                    existente.numero_processo = numero_processo[:100]
                    existente.save()
                else:
                    Compromisso.objects.create(
                        titulo=titulo[:200],
                        tipo='atendimento',
                        data=evento_data,
                        hora_inicio=hora_inicio,
                        hora_fim=hora_fim,
                        local=local[:200],
                        descricao='',
                        cor='#1d4ed8',
                        presencial=presencial,
                        numero_processo=numero_processo[:100],
                        criado_por=request.user,
                    )
                    importados += 1
            except Exception as e:
                erros.append(f"Linha {i}: {str(e)}")

    except Exception as e:
        return JsonResponse({'error': f'Erro ao processar arquivo: {str(e)}'}, status=400)

    return JsonResponse({'ok': True, 'importados': importados, 'erros': erros})


# ─── WEBHOOK POWER AUTOMATE ────────────────────────────────────────────────────

from django.views.decorators.csrf import csrf_exempt

# Token secreto configurado no settings.py ou hardcoded aqui.
# O Power Automate enviará esse token no header X-Webhook-Token.
# ALTERE PARA UM VALOR SEGURO E ÚNICO antes de usar em produção.
AGENDA_WEBHOOK_TOKEN = getattr(__import__('django.conf', fromlist=['settings']).settings,
                               'AGENDA_WEBHOOK_TOKEN', 'TROQUE_ESTE_TOKEN_SECRETO')


@csrf_exempt
def agenda_webhook_bookings(request):
    """
    Endpoint chamado automaticamente pelo Power Automate quando uma reserva
    é criada/atualizada no Microsoft Bookings.

    Autenticação: header  X-Webhook-Token: <token>

    Payload JSON esperado (enviado pelo Power Automate):
    {
        "customerName": "João Silva",
        "startDateTime": "2026-04-15T15:00:00",
        "endDateTime":   "2026-04-15T15:30:00",
        "serviceNotes":  "...",
        "customQuestionAnswers": [
            {"question": "Reunião Presencial?", "answer": "Sim"},
            {"question": "Número do Processo:", "answer": "1234567-89.2023.8.11.0001"}
        ]
    }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)

    # ── Autenticação por token ──
    token = request.headers.get('X-Webhook-Token', '')
    if token != AGENDA_WEBHOOK_TOKEN:
        return JsonResponse({'error': 'Token inválido'}, status=401)

    try:
        import json as _json
        from datetime import datetime as _dt

        body = _json.loads(request.body)

        # ── Nome / título ──
        titulo = str(body.get('customerName', '') or body.get('CustomerName', '') or 'Atendimento Bookings').strip()
        if not titulo:
            titulo = 'Atendimento Bookings'

        # ── Data e hora (ISO 8601: "2026-04-15T15:00:00" ou "2026-04-15T15:00:00Z") ──
        def _parse_iso(s):
            s = str(s or '').strip().replace('Z', '')
            for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
                try:
                    return _dt.strptime(s, fmt)
                except ValueError:
                    continue
            return None

        raw_start = body.get('startDateTime') or body.get('StartDateTime') or body.get('start') or ''
        raw_end   = body.get('endDateTime')   or body.get('EndDateTime')   or body.get('end')   or ''

        dt_start = _parse_iso(raw_start)
        dt_end   = _parse_iso(raw_end)

        if not dt_start:
            return JsonResponse({'error': f"startDateTime inválido: '{raw_start}'"}, status=400)

        evento_data   = dt_start.date()
        hora_inicio   = dt_start.time()
        hora_fim      = dt_end.time() if dt_end else None

        # ── Custom Question Answers (Reunião Presencial? / Número do Processo:) ──
        presencial = True
        numero_processo = ''

        answers = body.get('customQuestionAnswers') or body.get('CustomQuestionAnswers') or []
        for qa in answers:
            q = str(qa.get('question', '') or qa.get('Question', '')).strip()
            a = str(qa.get('answer',   '') or qa.get('Answer',   '')).strip()
            if 'presencial' in q.lower():
                if a.lower() in ('não', 'nao', 'no', 'false', '0', 'virtual'):
                    presencial = False
            if 'processo' in q.lower():
                numero_processo = a

        # Fallback plano (campos diretos sem aninhamento)
        if not answers:
            pres_raw = str(body.get('presencial', '') or body.get('locationtype', '')).lower()
            if pres_raw in ('não', 'nao', 'no', 'virtual', 'false', '0'):
                presencial = False
            numero_processo = str(body.get('processo', '') or body.get('processNumber', '')).strip()

        # ── Local e notas ──
        local    = str(body.get('location', '') or body.get('Location', '') or '').strip()[:200]
        descricao = str(body.get('serviceNotes', '') or body.get('notes', '') or '').strip()

        Compromisso.objects.create(
            titulo=titulo[:200],
            tipo='atendimento',
            data=evento_data,
            hora_inicio=hora_inicio,
            hora_fim=hora_fim,
            local=local,
            descricao=descricao,
            cor='#1d4ed8',
            presencial=presencial,
            numero_processo=numero_processo[:100],
            criado_por=None,  # vem de automação externa
        )

        return JsonResponse({'ok': True, 'mensagem': f'Compromisso criado: {titulo}'})

    except Exception as e:
        return JsonResponse({'error': f'Erro ao processar webhook: {str(e)}'}, status=400)

@login_required
def agenda_cancelar(request, pk):
    from django.shortcuts import get_object_or_404
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido.'}, status=405)
    if not _is_agenda_authorized(request):
        return JsonResponse({'error': 'Acesso negado.'}, status=403)
    
    comp = get_object_or_404(Compromisso, pk=pk)
    comp.cancelado = not comp.cancelado
    comp.save()
    status_msg = "cancelado" if comp.cancelado else "restaurado"
    return JsonResponse({'ok': True, 'mensagem': f'Compromisso {status_msg} com sucesso.', 'cancelado': comp.cancelado})
